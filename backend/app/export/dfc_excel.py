from __future__ import annotations

from copy import copy
from io import BytesIO

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.domain import DfcCalculation


SUMMARY_HEADERS = (
    "tipo_registro",
    "analise",
    "exercicio",
    "fca",
    "status_fca",
    "valor_automatico",
    "fluxo_operacional",
    "fluxo_investimento",
    "fluxo_financiamento",
    "ajustes_manuais_validados",
    "atividade",
    "codigo_componente",
    "componente",
    "valor_componente",
    "quantidade_movimentos",
    "codigo_pendencia",
    "mensagem_pendencia",
    "lancamento_pendencia",
    "linha_pendencia",
    "materialidade",
    "bloqueia_fca",
    "alertas",
    "limitacoes",
    "versao_metodologica",
    "sem_recalculo",
)

AUDIT_HEADERS = (
    "analise",
    "exercicio",
    "lancamento",
    "data_lancamento",
    "conta_caixa",
    "direcao_fluxo",
    "codigo_contrapartida",
    "nome_contrapartida",
    "codigo_referencial_contrapartida",
    "atividade_dfc",
    "codigo_componente",
    "componente",
    "valor_movimento",
    "valor_incluido",
    "status_final",
    "motivo_pendencia",
    "historico",
    "linha_origem",
    "status_fca",
    "versao_metodologica",
    "sem_recalculo",
)


def build_dfc_workbook(
    calculation: DfcCalculation,
    *,
    analysis_id: str,
) -> Workbook:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "dfc_resumo"
    summary.append(SUMMARY_HEADERS)
    summary.append(
        _summary_row(
            record_type="calculo",
            calculation=calculation,
            analysis_id=analysis_id,
        )
    )
    for component in calculation.component_summaries:
        summary.append(
            _summary_row(
                record_type="componente",
                calculation=calculation,
                analysis_id=analysis_id,
                activity=component.activity.value,
                component_code=component.component_code,
                component_label=component.component_label,
                component_value=format(component.value, "f"),
                movement_count=component.movement_count,
            )
        )
    for issue in calculation.pending_issues:
        summary.append(
            _summary_row(
                record_type="pendencia",
                calculation=calculation,
                analysis_id=analysis_id,
                issue_code=issue.code,
                issue_message=issue.message,
                issue_entry_number=issue.entry_number,
                issue_line_number=issue.line_number,
                materiality_level=issue.materiality_level,
                blocks_fca="true" if issue.blocks_fca else "false",
            )
        )

    audit = workbook.create_sheet("dfc_auditoria")
    audit.append(AUDIT_HEADERS)
    for row in calculation.audit_rows:
        audit.append(
            [
                analysis_id,
                calculation.exercise_year,
                row.entry_number,
                row.entry_date.isoformat() if row.entry_date else None,
                row.cash_account_code,
                row.cash_flow_direction.value,
                row.counterparty_account_code,
                row.counterparty_account_name,
                row.counterparty_reference_code,
                row.dfc_activity.value,
                row.dfc_component_code,
                row.dfc_component_label,
                format(row.movement_value, "f"),
                format(row.included_value, "f"),
                row.final_status.value,
                row.pending_reason,
                row.history,
                row.line_number,
                calculation.status.value,
                calculation.methodology_version_id,
                "true",
            ]
        )

    _style_sheet(summary)
    _style_sheet(audit)
    return workbook


def serialize_dfc_workbook(
    calculation: DfcCalculation,
    *,
    analysis_id: str,
) -> bytes:
    output = BytesIO()
    build_dfc_workbook(calculation, analysis_id=analysis_id).save(output)
    return output.getvalue()


def _summary_row(
    *,
    record_type: str,
    calculation: DfcCalculation,
    analysis_id: str,
    activity: str | None = None,
    component_code: str | None = None,
    component_label: str | None = None,
    component_value: str | None = None,
    movement_count: int | None = None,
    issue_code: str | None = None,
    issue_message: str | None = None,
    issue_entry_number: str | None = None,
    issue_line_number: int | None = None,
    materiality_level: str | None = None,
    blocks_fca: str | None = None,
) -> list[object | None]:
    is_calculation = record_type == "calculo"
    return [
        record_type,
        analysis_id,
        calculation.exercise_year,
        format(calculation.fca_value, "f") if is_calculation else None,
        calculation.status.value if is_calculation else None,
        format(calculation.automatic_value, "f") if is_calculation else None,
        format(calculation.operational_flow, "f") if is_calculation else None,
        format(calculation.investment_flow, "f") if is_calculation else None,
        format(calculation.financing_flow, "f") if is_calculation else None,
        format(calculation.manual_adjustments_value, "f")
        if is_calculation
        else None,
        activity,
        component_code,
        component_label,
        component_value,
        movement_count,
        issue_code,
        issue_message,
        issue_entry_number,
        issue_line_number,
        materiality_level,
        blocks_fca,
        "\n".join(calculation.alerts) if is_calculation else None,
        "\n".join(calculation.limitations) if is_calculation else None,
        calculation.methodology_version_id,
        "true",
    ]


def _style_sheet(sheet: Worksheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        width = min(max([len(value) for value in values] or [8]) + 2, 48)
        sheet.column_dimensions[column_cells[0].column_letter].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            alignment = copy(cell.alignment)
            alignment.wrap_text = True
            alignment.vertical = "top"
            cell.alignment = alignment
