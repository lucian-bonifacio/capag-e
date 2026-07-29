from __future__ import annotations

from copy import copy
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.domain import PlraCalculation


SUMMARY_HEADERS = [
    "analise",
    "exercicio",
    "ativos_brutos",
    "passivos_economicos_exigiveis",
    "ativos_ajustados",
    "plr_bruto",
    "plra",
    "status_plra",
    "formula",
    "contas_pendentes",
    "warnings",
    "limitacoes",
    "bloqueios",
    "status_balanco_declarado",
    "versao_metodologica",
    "calculado_em",
    "sem_recalculo",
]

MEMORY_HEADERS = [
    "codigo_conta",
    "nome_conta",
    "tipo_conta",
    "nivel_conta",
    "codigo_conta_pai",
    "cod_cta_ref_declarado",
    "descricao_oficial",
    "regra_metodologica",
    "grupo_metodologico",
    "macrogrupo",
    "valor_contabil",
    "sinal",
    "status_inclusao",
    "desagio_default",
    "valor_economico_default",
    "fonte_avaliacao",
    "valor_avaliacao_validada",
    "valor_economico_final",
    "status_decisao",
    "status_evidencia",
    "motivo",
    "limitacoes",
    "versao_metodologica",
    "sem_recalculo",
]


def build_plra_workbook(calculation: PlraCalculation) -> Workbook:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "plra_resumo"
    summary.append(SUMMARY_HEADERS)
    summary.append(
        [
            calculation.analysis_id,
            calculation.exercise_year,
            _decimal_string(calculation.gross_assets_value),
            _decimal_string(calculation.gross_economic_liabilities_value),
            _decimal_string(calculation.adjusted_assets_value),
            _decimal_string(calculation.plr_gross_value),
            _decimal_string(calculation.plra_value),
            calculation.plra_status.value,
            calculation.calculation_formula,
            _join_messages(calculation.pending_accounts),
            _join_messages(calculation.warnings),
            _join_messages(calculation.limitations),
            _join_messages(calculation.blocking_issues),
            calculation.balance_status.value,
            calculation.methodology_version_id,
            calculation.calculated_at.isoformat(),
            "true",
        ]
    )

    memory = workbook.create_sheet("plra_memoria")
    memory.append(MEMORY_HEADERS)
    for row in calculation.account_rows:
        memory.append(
            [
                row.account_code,
                row.account_name,
                row.account_type,
                row.account_level,
                row.parent_account_code,
                row.declared_reference_code,
                row.official_description,
                row.methodology_rule_id,
                row.methodology_group,
                row.macrogroup,
                _decimal_string(row.base_value),
                row.sign,
                row.inclusion_status.value,
                _optional_decimal_string(row.default_discount_percent),
                _decimal_string(row.default_economic_value),
                row.valuation_source,
                _optional_decimal_string(row.validated_valuation_value),
                _decimal_string(row.final_economic_value),
                row.decision_status.value,
                row.evidence_status,
                row.reason,
                _join_messages(row.limitations),
                row.methodology_version_id,
                "true",
            ]
        )

    _style_sheet(summary)
    _style_sheet(memory)
    return workbook


def serialize_plra_workbook(calculation: PlraCalculation) -> bytes:
    workbook = build_plra_workbook(calculation)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _decimal_string(value: Decimal) -> str:
    return format(value, "f")


def _optional_decimal_string(value: Decimal | None) -> str | None:
    return None if value is None else _decimal_string(value)


def _join_messages(messages: tuple[str, ...]) -> str:
    return "\n".join(messages)


def _style_sheet(sheet: Worksheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        width = min(max([len(value) for value in values] or [8]) + 2, 42)
        sheet.column_dimensions[column_cells[0].column_letter].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            alignment = copy(cell.alignment)
            alignment.wrap_text = True
            alignment.vertical = "top"
            cell.alignment = alignment
