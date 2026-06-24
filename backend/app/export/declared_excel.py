from __future__ import annotations

from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.application.declared_service import DeclaredAccountSnapshotView, DeclaredLayerSummary


RESULT_HEADERS = [
    "conta_societaria",
    "nome_conta",
    "cod_cta_ref_declarado",
    "descricao_oficial",
    "regra_metodologica_aplicada",
    "status_regra",
    "finalidade",
    "tratamento",
    "valor_base",
    "valor_considerado",
    "status_final",
    "observacao",
    "acao_recomendada",
    "versao_metodologica",
]

PENDING_STATUSES = {
    "NAO_MAPEADO_METODOLOGICAMENTE",
    "COD_CTA_REF_NAO_ENCONTRADO_NA_TABELA_OFICIAL",
    "SEM_VINCULO_REFERENCIAL",
    "REGRA_BLOQUEADA",
    "REGRA_EM_REVISAO",
    "REGRA_DEPRECIADA",
}


def build_declared_layer_workbook(
    *,
    summary: DeclaredLayerSummary,
    accounts: Iterable[DeclaredAccountSnapshotView],
) -> Workbook:
    account_list = list(accounts)
    workbook = Workbook()
    workbook.remove(workbook.active)

    _write_summary_sheet(workbook, summary)
    _write_result_sheet(workbook, "campos_resultado", account_list)
    _write_result_sheet(workbook, "memoria_calculo", account_list)
    _write_result_sheet(
        workbook,
        "fco_status",
        [account for account in account_list if account.purpose == "FCO"],
    )
    _write_result_sheet(
        workbook,
        "alertas_pendencias",
        [account for account in account_list if account.final_status in PENDING_STATUSES],
    )
    _write_audit_sheet(workbook, summary)
    _write_result_sheet(
        workbook,
        "inconsistencias",
        [
            account
            for account in account_list
            if account.final_status
            in {
                "NAO_MAPEADO_METODOLOGICAMENTE",
                "COD_CTA_REF_NAO_ENCONTRADO_NA_TABELA_OFICIAL",
            }
        ],
    )
    _write_result_sheet(
        workbook,
        "sem_vinculo_ref",
        [account for account in account_list if account.final_status == "SEM_VINCULO_REFERENCIAL"],
    )

    return workbook


def serialize_declared_layer_workbook(
    *,
    summary: DeclaredLayerSummary,
    accounts: Iterable[DeclaredAccountSnapshotView],
) -> bytes:
    workbook = build_declared_layer_workbook(summary=summary, accounts=accounts)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _write_summary_sheet(workbook: Workbook, summary: DeclaredLayerSummary) -> None:
    sheet = workbook.create_sheet("resumo_executivo")
    rows = [
        ("analysis_id", summary.analysis_id),
        ("exercicio", summary.year),
        ("total_contas", summary.total_accounts),
        ("versao_metodologica", summary.methodology_version_id),
    ]

    for row in rows:
        sheet.append(row)

    sheet.append(())
    sheet.append(("status_final", "quantidade"))
    for status, count in sorted(summary.status_counts.items()):
        sheet.append((status, count))

    _style_sheet(sheet)


def _write_result_sheet(
    workbook: Workbook,
    title: str,
    accounts: Iterable[DeclaredAccountSnapshotView],
) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(RESULT_HEADERS)

    for account in accounts:
        sheet.append(
            [
                account.account_code,
                account.account_name,
                account.declared_reference_code,
                account.official_description,
                account.methodology_rule_applied,
                account.methodology_rule_status,
                account.purpose,
                account.treatment,
                str(account.base_value),
                str(account.considered_value),
                account.final_status,
                account.observation,
                account.recommended_action,
                account.methodology_version_id,
            ]
        )

    _style_sheet(sheet)


def _write_audit_sheet(workbook: Workbook, summary: DeclaredLayerSummary) -> None:
    sheet = workbook.create_sheet("log_auditoria")
    sheet.append(("evento", "detalhe"))
    sheet.append(("fonte", "snapshot_declarado_persistido"))
    sheet.append(("sem_recalculo", "true"))
    sheet.append(("analysis_id", summary.analysis_id))
    sheet.append(("exercicio", summary.year))
    sheet.append(("versao_metodologica", summary.methodology_version_id))
    _style_sheet(sheet)


def _style_sheet(sheet: Worksheet) -> None:
    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        width = min(max([len(value) for value in values] or [8]) + 2, 42)
        sheet.column_dimensions[column_cells[0].column_letter].width = width
