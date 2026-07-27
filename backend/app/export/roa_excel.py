from __future__ import annotations

from copy import copy
from io import BytesIO

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.domain import CapagEAssessment, RoaBlock, RoaCalculation


SUMMARY_HEADERS = (
    "tipo_registro",
    "analise",
    "exercicio",
    "receita_bruta",
    "deducoes",
    "tributos_receita",
    "receita_operacional_liquida",
    "custos_operacionais",
    "despesas_operacionais",
    "resultado_financeiro",
    "resultado_nao_operacional",
    "pressoes_caixa",
    "roa_preliminar",
    "roa_final",
    "status_roa",
    "bloco_roa",
    "codigo_componente",
    "componente",
    "valor_componente",
    "quantidade_contas",
    "codigo_pendencia",
    "mensagem_pendencia",
    "conta_pendencia",
    "codigo_referencial_pendencia",
    "materialidade",
    "evidencia",
    "bloqueia_roa",
    "plra",
    "status_plra",
    "fca",
    "status_fca",
    "capag_e",
    "status_capag_e",
    "metodo_capag_e",
    "formula_capag_e",
    "base_calculo_capag_e",
    "alertas",
    "limitacoes",
    "versao_metodologica",
    "sem_recalculo",
)

AUDIT_HEADERS = (
    "analise",
    "exercicio",
    "codigo_conta",
    "nome_conta",
    "codigo_referencial",
    "descricao_referencial",
    "bloco_roa",
    "codigo_componente",
    "componente",
    "valor_base",
    "efeito_roa",
    "tratamento",
    "status_final",
    "motivo_pendencia",
    "evidencia",
    "linha_origem",
    "macrogrupo",
    "tipo_evidencia_exigida",
    "detalhe_fonte",
    "status_roa",
    "versao_metodologica",
    "sem_recalculo",
)

PRESSURE_HEADERS = (
    "analise",
    "exercicio",
    "codigo_pressao",
    "nome_pressao",
    "codigo_referencial",
    "codigo_componente",
    "componente",
    "valor_pressao",
    "efeito_roa",
    "status_final",
    "motivo_pendencia",
    "evidencia",
    "tipo_evidencia_exigida",
    "detalhe_fonte",
    "status_roa",
    "versao_metodologica",
    "sem_recalculo",
)


def build_roa_workbook(
    calculation: RoaCalculation,
    *,
    analysis_id: str,
    capag_assessment: CapagEAssessment | None = None,
) -> Workbook:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "roa_resumo"
    summary.append(SUMMARY_HEADERS)
    summary.append(
        _summary_row(
            record_type="calculo",
            calculation=calculation,
            analysis_id=analysis_id,
            capag_assessment=capag_assessment,
        )
    )
    for component in calculation.component_summaries:
        summary.append(
            _summary_row(
                record_type="componente",
                calculation=calculation,
                analysis_id=analysis_id,
                capag_assessment=capag_assessment,
                block=component.block.value,
                component_code=component.component_code,
                component_label=component.component_label,
                component_value=format(component.value, "f"),
                account_count=component.account_count,
            )
        )
    for group in calculation.pending_groups:
        summary.append(
            _summary_row(
                record_type="pendencia",
                calculation=calculation,
                analysis_id=analysis_id,
                capag_assessment=capag_assessment,
                pending_code=group.code,
                pending_message=group.message,
                pending_account=group.account_code,
                pending_reference=group.reference_code,
                materiality_level=group.materiality_level,
                evidence_id=group.evidence_id,
                blocks_roa="true" if group.blocks_roa else "false",
            )
        )

    audit = workbook.create_sheet("roa_auditoria")
    audit.append(AUDIT_HEADERS)
    for row in calculation.audit_rows:
        audit.append(
            [
                analysis_id,
                calculation.exercise_year,
                row.account_code,
                row.account_name,
                row.reference_code,
                row.reference_description,
                row.roa_block.value if row.roa_block else None,
                row.component_roa,
                row.component_label,
                format(row.base_value, "f"),
                format(row.signed_value, "f"),
                row.treatment,
                row.final_status.value,
                row.pending_reason,
                row.evidence_id,
                row.line_reference,
                row.macrogroup,
                row.required_evidence_type,
                row.source_detail,
                calculation.status.value,
                calculation.methodology_version_id,
                "true",
            ]
        )

    pressures = workbook.create_sheet("roa_pressoes_caixa")
    pressures.append(PRESSURE_HEADERS)
    for row in calculation.audit_rows:
        if row.roa_block != RoaBlock.CASH_PRESSURES:
            continue
        pressures.append(
            [
                analysis_id,
                calculation.exercise_year,
                row.account_code,
                row.account_name,
                row.reference_code,
                row.component_roa,
                row.component_label,
                format(row.base_value, "f"),
                format(row.signed_value, "f"),
                row.final_status.value,
                row.pending_reason,
                row.evidence_id,
                row.required_evidence_type,
                row.source_detail,
                calculation.status.value,
                calculation.methodology_version_id,
                "true",
            ]
        )

    for sheet in (summary, audit, pressures):
        _style_sheet(sheet)
    return workbook


def serialize_roa_workbook(
    calculation: RoaCalculation,
    *,
    analysis_id: str,
    capag_assessment: CapagEAssessment | None = None,
) -> bytes:
    output = BytesIO()
    build_roa_workbook(
        calculation,
        analysis_id=analysis_id,
        capag_assessment=capag_assessment,
    ).save(output)
    return output.getvalue()


def _summary_row(
    *,
    record_type: str,
    calculation: RoaCalculation,
    analysis_id: str,
    capag_assessment: CapagEAssessment | None,
    block: str | None = None,
    component_code: str | None = None,
    component_label: str | None = None,
    component_value: str | None = None,
    account_count: int | None = None,
    pending_code: str | None = None,
    pending_message: str | None = None,
    pending_account: str | None = None,
    pending_reference: str | None = None,
    materiality_level: str | None = None,
    evidence_id: str | None = None,
    blocks_roa: str | None = None,
) -> list[object | None]:
    is_calculation = record_type == "calculo"
    assessment = capag_assessment if is_calculation else None
    return [
        record_type,
        analysis_id,
        calculation.exercise_year,
        _calculation_value(calculation.gross_revenue, is_calculation),
        _calculation_value(calculation.deductions, is_calculation),
        _calculation_value(calculation.revenue_taxes, is_calculation),
        _calculation_value(calculation.net_operating_revenue, is_calculation),
        _calculation_value(calculation.operating_costs, is_calculation),
        _calculation_value(calculation.operating_expenses, is_calculation),
        _calculation_value(calculation.financial_result, is_calculation),
        _calculation_value(calculation.non_operating_result, is_calculation),
        _calculation_value(calculation.cash_pressure_adjustments, is_calculation),
        _calculation_value(calculation.roa_preliminary, is_calculation),
        _calculation_value(calculation.roa_final, is_calculation),
        calculation.status.value if is_calculation else None,
        block,
        component_code,
        component_label,
        component_value,
        account_count,
        pending_code,
        pending_message,
        pending_account,
        pending_reference,
        materiality_level,
        evidence_id,
        blocks_roa,
        _assessment_value(assessment, "plra_value"),
        assessment.plra_status.value if assessment else None,
        _assessment_value(assessment, "fca_value"),
        assessment.fca_status.value if assessment else None,
        _assessment_value(assessment, "capag_e_value"),
        assessment.capag_e_status.value if assessment else None,
        assessment.method.value if assessment else None,
        assessment.methodology_formula if assessment else None,
        assessment.calculation_basis if assessment else None,
        "\n".join(calculation.alerts) if is_calculation else None,
        "\n".join(calculation.limitations) if is_calculation else None,
        calculation.methodology_version_id,
        "true",
    ]


def _calculation_value(value, enabled: bool) -> str | None:
    return format(value, "f") if enabled else None


def _assessment_value(
    assessment: CapagEAssessment | None,
    field_name: str,
) -> str | None:
    if assessment is None:
        return None
    value = getattr(assessment, field_name)
    return None if value is None else format(value, "f")


def _style_sheet(sheet: Worksheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        width = min(max([len(value) for value in values] or [8]) + 2, 48)
        sheet.column_dimensions[column_cells[0].column_letter].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            alignment = copy(cell.alignment)
            alignment.wrap_text = True
            alignment.vertical = "top"
            cell.alignment = alignment
