from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.domain import AdjustmentEvidence, AssetValuationAssessment


EVIDENCE_HEADERS = (
    "evidence_id",
    "exercise_year",
    "scope_type",
    "scope_key",
    "adjustment_type",
    "method_component",
    "amount_impact",
    "impact_base_value",
    "impact_percent",
    "materiality_level",
    "materiality_source",
    "minimum_materiality_level",
    "required_evidence_type",
    "evidence_status",
    "analyst_justification",
    "review_notes",
    "blocks_final_report",
    "requires_reservation",
    "human_review_required",
    "decision_reasons",
    "materiality_overrides",
    "methodology_version_id",
)

ASSET_HEADERS = (
    "assessment_id",
    "exercise_year",
    "account_code",
    "account_name",
    "reference_code",
    "macrogroup",
    "book_value",
    "default_desagio_percent",
    "default_economic_value",
    "valuation_required",
    "realizability_classification",
    "valuation_basis",
    "forced_liquidation_value",
    "analyst_adjusted_value",
    "final_economic_value",
    "final_value_source",
    "essentiality_status",
    "evidence_id",
    "valuation_status",
    "blocks_plra",
    "blocking_reasons",
    "methodology_version_id",
)


def build_evidence_workbook(
    evidences: list[AdjustmentEvidence],
    asset_valuations: list[AssetValuationAssessment],
) -> Workbook:
    workbook = Workbook()
    evidence_sheet = workbook.active
    evidence_sheet.title = "evidencias_justificativas"
    asset_sheet = workbook.create_sheet("avaliacao_ativos")

    _write_sheet(
        evidence_sheet,
        EVIDENCE_HEADERS,
        [_evidence_row(evidence) for evidence in evidences],
    )
    _write_sheet(
        asset_sheet,
        ASSET_HEADERS,
        [_asset_row(assessment) for assessment in asset_valuations],
    )
    return workbook


def serialize_evidence_workbook(
    evidences: list[AdjustmentEvidence],
    asset_valuations: list[AssetValuationAssessment],
) -> bytes:
    output = BytesIO()
    build_evidence_workbook(evidences, asset_valuations).save(output)
    return output.getvalue()


def _evidence_row(evidence: AdjustmentEvidence) -> list[object]:
    snapshot = evidence.to_snapshot()
    snapshot["decision_reasons"] = " | ".join(evidence.decision_reasons)
    snapshot["materiality_overrides"] = " | ".join(
        f"{item.before.value}->{item.after.value}: {item.justification}"
        for item in evidence.materiality_overrides
    )
    return [snapshot.get(header) for header in EVIDENCE_HEADERS]


def _asset_row(assessment: AssetValuationAssessment) -> list[object]:
    snapshot = assessment.to_snapshot()
    snapshot["blocking_reasons"] = " | ".join(assessment.blocking_reasons)
    return [snapshot.get(header) for header in ASSET_HEADERS]


def _write_sheet(sheet, headers: tuple[str, ...], rows: list[list[object]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="334155")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        width = min(
            max(len(str(cell.value or "")) for cell in column) + 2,
            48,
        )
        sheet.column_dimensions[column[0].column_letter].width = max(width, 12)
