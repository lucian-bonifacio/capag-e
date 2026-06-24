from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.application.declared_service import DeclaredAccountSnapshotView, DeclaredLayerSummary
from app.export import build_declared_layer_workbook, serialize_declared_layer_workbook


def test_declared_excel_export_serializes_snapshot_fields_without_recalculation() -> None:
    summary = _summary()
    accounts = [
        _account(
            account_code="1725",
            declared_reference_code="2.01.01.07.01",
            methodology_rule_applied="2.01.01.07.01",
            base_value=Decimal("100000.00"),
            considered_value=Decimal("0.00"),
            final_status="MAPEADO",
        )
    ]

    workbook = build_declared_layer_workbook(summary=summary, accounts=accounts)

    assert workbook.sheetnames == [
        "resumo_executivo",
        "campos_resultado",
        "memoria_calculo",
        "fco_status",
        "alertas_pendencias",
        "log_auditoria",
        "inconsistencias",
        "sem_vinculo_ref",
    ]
    result_sheet = workbook["campos_resultado"]
    assert result_sheet["A2"].value == "1725"
    assert result_sheet["C2"].value == "2.01.01.07.01"
    assert result_sheet["E2"].value == "2.01.01.07.01"
    assert result_sheet["I2"].value == "100000.00"
    assert result_sheet["J2"].value == "0.00"
    assert result_sheet["K2"].value == "MAPEADO"
    assert result_sheet["N2"].value == "test-version"
    assert _has_no_formulas(workbook)


def test_declared_excel_export_preserves_dangerous_prefix_case_as_not_mapped() -> None:
    workbook = build_declared_layer_workbook(
        summary=_summary(status_counts={"NAO_MAPEADO_METODOLOGICAMENTE": 1}),
        accounts=[
            _account(
                account_code="1725",
                declared_reference_code="2.01.01.07.01",
                methodology_rule_applied=None,
                final_status="NAO_MAPEADO_METODOLOGICAMENTE",
                observation="Nenhuma regra metodologica exata encontrada para a finalidade.",
                recommended_action="revisar_metodologia",
            )
        ],
    )

    result_sheet = workbook["campos_resultado"]
    pending_sheet = workbook["alertas_pendencias"]

    assert result_sheet["C2"].value == "2.01.01.07.01"
    assert result_sheet["E2"].value is None
    assert result_sheet["K2"].value == "NAO_MAPEADO_METODOLOGICAMENTE"
    assert result_sheet["M2"].value == "revisar_metodologia"
    assert pending_sheet["K2"].value == "NAO_MAPEADO_METODOLOGICAMENTE"
    assert _has_no_formulas(workbook)


def test_declared_excel_export_returns_xlsx_bytes_readable_by_openpyxl() -> None:
    payload = serialize_declared_layer_workbook(
        summary=_summary(),
        accounts=[
            _account(
                account_code="3001",
                declared_reference_code=None,
                methodology_rule_applied=None,
                final_status="SEM_VINCULO_REFERENCIAL",
            )
        ],
    )

    workbook = load_workbook(BytesIO(payload))

    assert workbook["sem_vinculo_ref"]["A2"].value == "3001"
    assert workbook["sem_vinculo_ref"]["K2"].value == "SEM_VINCULO_REFERENCIAL"
    assert workbook["log_auditoria"]["B3"].value == "true"
    assert _has_no_formulas(workbook)


def _summary(status_counts: dict[str, int] | None = None) -> DeclaredLayerSummary:
    return DeclaredLayerSummary(
        analysis_id="analysis-1",
        year=2024,
        total_accounts=1,
        status_counts=status_counts or {"MAPEADO": 1},
        methodology_version_id="test-version",
    )


def _account(
    *,
    account_code: str,
    declared_reference_code: str | None,
    methodology_rule_applied: str | None,
    final_status: str,
    base_value: Decimal = Decimal("100.00"),
    considered_value: Decimal = Decimal("100.00"),
    observation: str | None = None,
    recommended_action: str | None = None,
) -> DeclaredAccountSnapshotView:
    return DeclaredAccountSnapshotView(
        account_code=account_code,
        account_name="Conta de teste",
        declared_reference_code=declared_reference_code,
        official_description="Descricao oficial" if declared_reference_code else None,
        official_reference_status="ATIVA" if declared_reference_code else None,
        methodology_rule_applied=methodology_rule_applied,
        methodology_rule_status="ATIVA" if methodology_rule_applied else None,
        purpose="FCO",
        treatment="tratamento_serializado" if methodology_rule_applied else None,
        base_value=base_value,
        considered_value=considered_value,
        final_status=final_status,
        observation=observation,
        recommended_action=recommended_action,
        methodology_version_id="test-version",
    )


def _has_no_formulas(workbook) -> bool:
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    return False

    return True
