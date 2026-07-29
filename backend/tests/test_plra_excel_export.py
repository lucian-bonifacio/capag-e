from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.domain import (
    ComponentStatus,
    PlraAccountAuditRow,
    PlraCalculation,
    PlraDecisionStatus,
    PlraInclusionStatus,
)
from app.export import build_plra_workbook, serialize_plra_workbook


def test_plra_excel_exports_summary_memory_defaults_overrides_and_pending_data() -> None:
    workbook = build_plra_workbook(_calculation())
    summary = _row_dict(workbook["plra_resumo"])
    memory = workbook["plra_memoria"]
    memory_headers = [cell.value for cell in memory[1]]
    first_row = dict(zip(memory_headers, (cell.value for cell in memory[2])))
    second_row = dict(zip(memory_headers, (cell.value for cell in memory[3])))

    assert workbook.sheetnames == ["plra_resumo", "plra_memoria"]
    assert summary["plra"] == "123.45"
    assert summary["status_plra"] == "bloqueado_por_pendencia"
    assert summary["contas_pendentes"] == "conditional"
    assert summary["bloqueios"] == "PASSIVO_CONDICIONAL_SEM_DECISAO:conditional"
    assert summary["status_balanco_declarado"] == "VALIDO"
    assert summary["versao_metodologica"] == "metodologia-2024.1"
    assert summary["sem_recalculo"] == "true"

    assert first_row["valor_contabil"] == "100.00"
    assert first_row["desagio_default"] == "0.30"
    assert first_row["valor_economico_default"] == "70.00"
    assert first_row["valor_avaliacao_validada"] == "60.00"
    assert first_row["valor_economico_final"] == "60.00"
    assert first_row["fonte_avaliacao"] == "avaliacao_validada"
    assert second_row["status_inclusao"] == "pendente"
    assert second_row["status_decisao"] == "pendente"
    assert _has_no_formulas(workbook)


def test_plra_excel_serializes_snapshot_values_without_recalculation() -> None:
    workbook = build_plra_workbook(_calculation())
    summary = _row_dict(workbook["plra_resumo"])

    assert summary["ativos_ajustados"] == "700.00"
    assert summary["passivos_economicos_exigiveis"] == "200.00"
    assert summary["plra"] == "123.45"
    assert summary["plra"] != "500.00"
    assert _has_no_formulas(workbook)


def test_plra_excel_returns_readable_xlsx_bytes() -> None:
    payload = serialize_plra_workbook(_calculation())
    workbook = load_workbook(BytesIO(payload))

    assert workbook["plra_resumo"]["G2"].value == "123.45"
    assert workbook["plra_memoria"]["R2"].value == "60.00"
    assert _has_no_formulas(workbook)


def _calculation() -> PlraCalculation:
    return PlraCalculation(
        analysis_id="analysis",
        exercise_year=2024,
        gross_assets_value=Decimal("1000"),
        gross_economic_liabilities_value=Decimal("200"),
        adjusted_assets_value=Decimal("700"),
        plr_gross_value=Decimal("800"),
        plra_value=Decimal("123.45"),
        plra_status=ComponentStatus.BLOCKED_BY_PENDING,
        calculation_formula="PLRA = ativos ajustados - passivos exigiveis",
        account_rows=(
            PlraAccountAuditRow(
                account_code="clients",
                account_name="Clientes",
                account_type="A",
                account_level=5,
                parent_account_code=None,
                declared_reference_code="1.01.02.02.01",
                official_description="Duplicatas a receber",
                methodology_rule_id="PLRA-clients",
                methodology_group="clientes",
                macrogroup="ATIVO_REALIZAVEL",
                base_value=Decimal("100"),
                sign="D",
                inclusion_status=PlraInclusionStatus.INCLUDED_ASSET,
                default_discount_percent=Decimal("0.30"),
                default_economic_value=Decimal("70"),
                valuation_source="avaliacao_validada",
                validated_valuation_value=Decimal("60"),
                final_economic_value=Decimal("60"),
                decision_status=PlraDecisionStatus.VALIDATED,
                evidence_status="validada",
                reason="Avaliacao validada substituiu o default.",
                limitations=(),
                methodology_version_id="metodologia-2024.1",
            ),
            PlraAccountAuditRow(
                account_code="conditional",
                account_name="Passivo condicional",
                account_type="A",
                account_level=5,
                parent_account_code=None,
                declared_reference_code="2.01.99",
                official_description="Outros passivos",
                methodology_rule_id="PLRA-conditional",
                methodology_group="passivo_condicional",
                macrogroup="PASSIVO_EXIGIVEL",
                base_value=Decimal("10"),
                sign="C",
                inclusion_status=PlraInclusionStatus.PENDING,
                default_discount_percent=None,
                default_economic_value=Decimal("0"),
                valuation_source=None,
                validated_valuation_value=None,
                final_economic_value=Decimal("0"),
                decision_status=PlraDecisionStatus.PENDING,
                evidence_status=None,
                reason="Decisao pendente.",
                limitations=("Revisao humana necessaria.",),
                methodology_version_id="metodologia-2024.1",
            ),
        ),
        pending_accounts=("conditional",),
        warnings=("J100 apenas informativo.",),
        limitations=("Cobertura parcial.",),
        blocking_issues=("PASSIVO_CONDICIONAL_SEM_DECISAO:conditional",),
        balance_status="VALIDO",
        methodology_version_id="metodologia-2024.1",
        calculated_at=datetime(2024, 12, 31, tzinfo=timezone.utc),
    )


def _row_dict(sheet) -> dict[str, object]:
    return dict(zip((cell.value for cell in sheet[1]), (cell.value for cell in sheet[2])))


def _has_no_formulas(workbook) -> bool:
    return all(
        not (isinstance(cell.value, str) and cell.value.startswith("="))
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )
