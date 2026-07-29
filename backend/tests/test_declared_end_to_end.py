from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.api.declared import (
    get_declared_run_session,
    get_declared_snapshot_reader,
    get_methodology_rules,
    get_official_references,
)
from app.api.imports import get_import_session
from app.application.declared_service import SqlAlchemyDeclaredSnapshotReader
from app.engine.methodology_matcher import MethodologyRule, OfficialReferenceAccount, RuleStatus
from app.main import app
from app.repositories import Base


FIXTURES_DIR = Path(__file__).parent / "fixtures" / "ecd"


@pytest.mark.parametrize(
    ("fixture_name", "expected_status", "official_references", "methodology_rules"),
    [
        (
            "valid_declared.ecd",
            "MAPEADO",
            ["2.01.01.07.01"],
            [("2.01.01.07.01", RuleStatus.ACTIVE)],
        ),
        ("missing_i051.ecd", "SEM_VINCULO_REFERENCIAL", ["2.01.01.07.01"], []),
        (
            "official_reference_missing.ecd",
            "COD_CTA_REF_NAO_ENCONTRADO_NA_TABELA_OFICIAL",
            ["2.01.01.07.01"],
            [("9.99.99.99.99", RuleStatus.ACTIVE)],
        ),
        (
            "methodology_missing.ecd",
            "NAO_MAPEADO_METODOLOGICAMENTE",
            ["1.01.02.03.04"],
            [],
        ),
        (
            "blocked_rule.ecd",
            "REGRA_BLOQUEADA",
            ["2.99.99.99.99"],
            [("2.99.99.99.99", RuleStatus.BLOCKED)],
        ),
        (
            "dangerous_prefix.ecd",
            "NAO_MAPEADO_METODOLOGICAMENTE",
            ["2.01.01.07.01"],
            [("2.01.01.*", RuleStatus.ACTIVE)],
        ),
    ],
)
def test_declared_layer_end_to_end_contract_for_governed_ecd_fixtures(
    fixture_name: str,
    expected_status: str,
    official_references: list[str],
    methodology_rules: list[tuple[str, RuleStatus]],
) -> None:
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    SessionForTest = sessionmaker(bind=engine)

    def override_session():
        session = SessionForTest()
        try:
            yield session
        finally:
            session.close()

    def override_reader():
        session = SessionForTest()
        try:
            yield SqlAlchemyDeclaredSnapshotReader(session)
        finally:
            session.close()

    app.dependency_overrides[get_import_session] = override_session
    app.dependency_overrides[get_declared_run_session] = override_session
    app.dependency_overrides[get_declared_snapshot_reader] = override_reader
    app.dependency_overrides[get_official_references] = lambda: [
        _official(reference_code) for reference_code in official_references
    ]
    app.dependency_overrides[get_methodology_rules] = lambda: [
        _rule(reference_code, rule_status) for reference_code, rule_status in methodology_rules
    ]
    client = TestClient(app)

    try:
        content = (FIXTURES_DIR / fixture_name).read_bytes()
        import_response = client.post(
            "/api/v1/ecd/import",
            data={"methodology_version_id": "metodologia-2024.1"},
            files={"file": (fixture_name, content, "text/plain")},
        )
        assert import_response.status_code == 201
        imported = import_response.json()

        run_response = client.post(
            f"/api/v1/analyses/{imported['analysis_id']}/exercises/{imported['year']}/declared/run"
        )
        assert run_response.status_code == 200
        assert run_response.json()["status_counts"] == {expected_status: 1}

        summary_response = client.get(
            f"/api/v1/analyses/{imported['analysis_id']}/exercises/{imported['year']}/declared"
        )
        assert summary_response.status_code == 200
        assert summary_response.json()["status_counts"] == {expected_status: 1}

        accounts_response = client.get(
            f"/api/v1/analyses/{imported['analysis_id']}/exercises/{imported['year']}/declared/accounts"
        )
        assert accounts_response.status_code == 200
        accounts = accounts_response.json()["accounts"]
        assert accounts[0]["final_status"] == expected_status
        assert "account_type" in accounts[0]
        assert "account_nature" in accounts[0]
        assert "account_level" in accounts[0]
        assert "parent_account_code" in accounts[0]
        assert "account_order" in accounts[0]

        balance_accounts_response = client.get(
            f"/api/v1/analyses/{imported['analysis_id']}/exercises/{imported['year']}/declared/balance/accounts"
        )
        assert balance_accounts_response.status_code == 200
        balance_payload = balance_accounts_response.json()
        assert "accounts" not in balance_payload
        assert balance_payload["balance_status"] == "ESTRUTURA_INVALIDA"
        assert balance_payload["is_blocking"] is True
        assert balance_payload["rows"] == []
        assert balance_payload["limitations"] == ["I010_AUSENTE_OU_AMBIGUO"]

        excel_response = client.get(
            f"/api/v1/analyses/{imported['analysis_id']}/exercises/{imported['year']}/declared/export.xlsx"
        )
        assert excel_response.status_code == 200
        workbook = load_workbook(BytesIO(excel_response.content))
        assert workbook["campos_resultado"]["K2"].value == expected_status
        assert _has_no_formulas(workbook)
    finally:
        app.dependency_overrides.clear()


def _official(reference_code: str) -> OfficialReferenceAccount:
    return OfficialReferenceAccount(
        reference_code=reference_code,
        official_description="Descricao oficial sintetica",
        parent_reference_code=None,
        level=5,
        nature="PASSIVO",
        valid_from=2020,
        valid_to=None,
        layout="ECD_2024",
        entity_type="PJ_GERAL",
        source="fixture_sintetica",
        status="ATIVA",
        methodology_version_id="metodologia-2024.1",
    )


def _rule(reference_code: str, rule_status: RuleStatus) -> MethodologyRule:
    return MethodologyRule(
        reference_code=reference_code,
        purpose="FCO",
        methodology_description="Regra sintetica para teste.",
        plra_category=None,
        fco_category="categoria_sintetica",
        capag_category=None,
        flow_nature=None,
        operational_treatment="tratamento_sintetico",
        include_in_calculation=True,
        sign=None,
        rule_status=rule_status,
        valid_from=2020,
        valid_to=None,
        methodology_version_id="metodologia-2024.1",
        observation=None,
    )


def _has_no_formulas(workbook) -> bool:
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    return False

    return True
