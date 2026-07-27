from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.assets.methodology import load_plra_policy
from app.domain import (
    AssetRealizability,
    EssentialityStatus,
    EvidenceScopeType,
    EvidenceStatus,
    MaterialityLevel,
    MethodComponent,
    ValuationBasis,
    ValuationStatus,
)
from app.engine import (
    apply_materiality_override,
    assess_asset_valuation,
    build_adjustment_evidence,
)
from app.export import build_evidence_workbook, serialize_evidence_workbook


def test_workbook_exports_critical_evidence_and_asset_audit_values() -> None:
    evidence = _evidence(status=EvidenceStatus.PENDING)
    valuation = _valuation(evidence=None)

    workbook = build_evidence_workbook([evidence], [valuation])

    assert workbook.sheetnames == [
        "evidencias_justificativas",
        "avaliacao_ativos",
    ]
    evidence_values = _row_by_header(
        workbook["evidencias_justificativas"]
    )
    assert evidence_values["amount_impact"] == "100.00"
    assert evidence_values["impact_percent"] == "0.100000"
    assert evidence_values["materiality_level"] == "critica"
    assert evidence_values["evidence_status"] == "pendente"
    assert evidence_values["blocks_final_report"] is True
    asset_values = _row_by_header(workbook["avaliacao_ativos"])
    assert asset_values["book_value"] == "1000.00"
    assert asset_values["default_desagio_percent"] == "0.800000"
    assert asset_values["default_economic_value"] == "200.00"
    assert asset_values["final_economic_value"] == "200.00"
    assert asset_values["valuation_status"] == "pendente"
    assert asset_values["blocks_plra"] is True


def test_workbook_preserves_override_history_without_recalculation() -> None:
    evidence = apply_materiality_override(
        _evidence(
            amount_impact=Decimal("20.00"),
            status=EvidenceStatus.VALIDATED,
        ),
        materiality_level=MaterialityLevel.HIGH,
        justification="Risco operacional confirmado.",
        overridden_at=datetime(2026, 7, 24, tzinfo=timezone.utc),
    )

    workbook = build_evidence_workbook([evidence], [])
    values = _row_by_header(workbook["evidencias_justificativas"])

    assert values["amount_impact"] == "20.00"
    assert values["impact_percent"] == "0.020000"
    assert values["materiality_level"] == "alta"
    assert (
        values["materiality_overrides"]
        == "media->alta: Risco operacional confirmado."
    )


def test_serialized_workbook_is_readable_and_has_no_formulas() -> None:
    payload = serialize_evidence_workbook(
        [_evidence(status=EvidenceStatus.PENDING)],
        [_valuation(evidence=None)],
    )

    workbook = load_workbook(BytesIO(payload), data_only=False)

    assert len(payload) > 1000
    assert all(
        not (isinstance(cell.value, str) and cell.value.startswith("="))
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )


def _evidence(
    *,
    amount_impact: Decimal = Decimal("100.00"),
    status: EvidenceStatus,
):
    return build_adjustment_evidence(
        evidence_id="evidence-export-1",
        exercise_year=2024,
        scope_type=EvidenceScopeType.ACCOUNT,
        scope_key="asset-1",
        adjustment_type="avaliacao_ativo",
        method_component=MethodComponent.PLRA,
        amount_impact=amount_impact,
        impact_base_value=Decimal("1000.00"),
        required_evidence_type="laudo_abnt_nbr_14653",
        evidence_status=status,
        analyst_justification="Avaliação patrimonial em revisão.",
        review_notes=None,
        methodology_version_id="metodologia-2024.1",
        created_at=datetime(2026, 7, 24, tzinfo=timezone.utc),
    )


def _valuation(*, evidence):
    return assess_asset_valuation(
        assessment_id="valuation-export-1",
        exercise_year=2024,
        account_code="asset-1",
        account_name="Maquinas e equipamentos",
        reference_code="1.02.03.01.06",
        book_value=Decimal("1000.00"),
        policy=load_plra_policy(),
        realizability_classification=(
            AssetRealizability.FORCED_LIQUIDATION_REQUIRES_REPORT
        ),
        valuation_required=True,
        valuation_basis=ValuationBasis.ABNT_NBR_14653_REPORT,
        forced_liquidation_value=Decimal("450.00"),
        analyst_adjusted_value=None,
        essentiality_status=EssentialityStatus.NOT_ESSENTIAL,
        valuation_status=ValuationStatus.PENDING,
        evidence=evidence,
    )


def _row_by_header(sheet) -> dict[str, object]:
    headers = [cell.value for cell in sheet[1]]
    return {
        header: sheet.cell(row=2, column=index + 1).value
        for index, header in enumerate(headers)
    }
