from datetime import date
from decimal import Decimal
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.api.evidence import get_evidence_session
from app.domain import ProcessingStatus
from app.main import create_app
from app.repositories import (
    AnalysisModel,
    Base,
    CompanyModel,
    EcdFileModel,
    EcdI050AccountModel,
    EcdI051ReferenceLinkModel,
    EcdI155BalanceModel,
    ExerciseModel,
)


EVIDENCES_PATH = "/api/v1/analyses/analysis-1/exercises/2024/evidences"
ASSETS_PATH = (
    "/api/v1/analyses/analysis-1/exercises/2024/assets/valuations"
)


def test_evidence_api_calculates_materiality_and_exposes_component_summary() -> None:
    client = _client()

    created = client.post(
        EVIDENCES_PATH,
        json=_evidence_payload(
            amount_impact="100.00",
            evidence_status="pendente",
        ),
    )

    assert created.status_code == 201
    body = created.json()
    assert body["amount_impact"] == "100.00"
    assert body["impact_base_value"] == "1000.00"
    assert body["impact_percent"] == "0.100000"
    assert body["materiality_level"] == "critica"
    assert body["blocks_final_report"] is True

    listed = client.get(EVIDENCES_PATH)
    assert listed.status_code == 200
    assert listed.json()["items"] == [body]
    assert listed.json()["summaries"] == [
        {
            "method_component": "PLRA",
            "total": 1,
            "blocking": 1,
            "reservations": 0,
            "pending": 1,
        }
    ]


def test_evidence_api_filters_and_records_justified_override() -> None:
    client = _client()
    created = client.post(EVIDENCES_PATH, json=_evidence_payload()).json()

    response = client.put(
        f"/api/v1/evidences/{created['evidence_id']}",
        json={
            "required_evidence_type": "documento_suporte",
            "evidence_status": "validada",
            "analyst_justification": "Conciliação validada.",
            "review_notes": "Revisado.",
            "materiality_override": {
                "materiality_level": "alta",
                "justification": "Risco operacional confirmado.",
            },
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["materiality_source"] == "override_manual"
    assert body["materiality_level"] == "alta"
    assert body["materiality_overrides"][0]["before"] == "media"
    assert body["materiality_overrides"][0]["after"] == "alta"

    filtered = client.get(
        EVIDENCES_PATH,
        params={"method_component": "PLRA", "evidence_status": "validada"},
    )
    assert filtered.status_code == 200
    assert len(filtered.json()["items"]) == 1


def test_evidence_api_rejects_waiver_or_override_without_justification() -> None:
    client = _client()

    waived = client.post(
        EVIDENCES_PATH,
        json=_evidence_payload(
            evidence_status="dispensada_com_justificativa",
            analyst_justification=None,
        ),
    )
    assert waived.status_code == 422

    created = client.post(EVIDENCES_PATH, json=_evidence_payload()).json()
    override = client.put(
        f"/api/v1/evidences/{created['evidence_id']}",
        json={
            "required_evidence_type": "documento_suporte",
            "evidence_status": "validada",
            "analyst_justification": "Conciliação validada.",
            "review_notes": None,
            "materiality_override": {
                "materiality_level": "alta",
                "justification": " ",
            },
        },
    )
    assert override.status_code == 422
    assert (
        override.json()["detail"]["error_code"]
        == "EVIDENCE_CONTRACT_ERROR"
    )


def test_evidence_api_rejects_numeric_decimal_payload() -> None:
    payload = _evidence_payload()
    payload["amount_impact"] = 20.0

    response = _client().post(EVIDENCES_PATH, json=payload)

    assert response.status_code == 422


def test_asset_api_applies_validated_value_and_serializes_decimal_strings() -> None:
    client = _client()
    evidence = client.post(
        EVIDENCES_PATH,
        json=_evidence_payload(
            amount_impact="250.00",
            evidence_status="validada",
        ),
    ).json()

    response = client.put(
        "/api/v1/assets/valuations/valuation-1",
        json=_asset_payload(evidence_id=evidence["evidence_id"]),
    )

    assert response.status_code == 200
    body = response.json()
    assert body["book_value"] == "1000.00"
    assert body["default_desagio_percent"] == "0.800000"
    assert body["default_economic_value"] == "200.00"
    assert body["forced_liquidation_value"] == "450.00"
    assert body["final_economic_value"] == "450.00"
    assert body["final_value_source"] == "liquidacao_forcada_validada"
    assert body["blocks_plra"] is False

    listed = client.get(ASSETS_PATH)
    assert listed.status_code == 200
    assert listed.json()["blocking_count"] == 0
    assert listed.json()["items"] == [body]


def test_asset_api_keeps_default_and_exposes_block_without_evidence() -> None:
    payload = _asset_payload(evidence_id=None)

    response = _client().put(
        "/api/v1/assets/valuations/valuation-1",
        json=payload,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["final_economic_value"] == "200.00"
    assert body["blocks_plra"] is True
    assert (
        "LIQUIDACAO_FORCADA_SEM_VALIDACAO_DOCUMENTAL"
        in body["blocking_reasons"]
    )


def test_evidence_api_downloads_both_governed_sheets() -> None:
    client = _client()
    evidence = client.post(
        EVIDENCES_PATH,
        json=_evidence_payload(
            amount_impact="100.00",
            evidence_status="pendente",
        ),
    ).json()
    client.put(
        "/api/v1/assets/valuations/valuation-1",
        json=_asset_payload(evidence_id=evidence["evidence_id"]),
    )

    response = client.get(f"{EVIDENCES_PATH}/export.xlsx")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "capag-evidencias-analysis-1-2024.xlsx" in response.headers[
        "content-disposition"
    ]
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == [
        "evidencias_justificativas",
        "avaliacao_ativos",
    ]
    assert (
        workbook["evidencias_justificativas"]["J2"].value == "critica"
    )
    assert workbook["avaliacao_ativos"]["S2"].value == "validada"


def test_evidence_openapi_exposes_governed_routes_and_string_decimals() -> None:
    schema = create_app().openapi()

    assert {"get", "post"} <= set(
        schema["paths"][
            "/api/v1/analyses/{analysis_id}/exercises/{year}/evidences"
        ]
    )
    assert "put" in schema["paths"]["/api/v1/evidences/{evidence_id}"]
    assert "get" in schema["paths"][
        "/api/v1/analyses/{analysis_id}/exercises/{year}/assets/valuations"
    ]
    assert "put" in schema["paths"][
        "/api/v1/assets/valuations/{assessment_id}"
    ]
    assert "get" in schema["paths"][
        "/api/v1/analyses/{analysis_id}/exercises/{year}/evidences/export.xlsx"
    ]
    evidence_schema = schema["components"]["schemas"]["EvidenceResponse"]
    assert evidence_schema["properties"]["amount_impact"]["type"] == "string"
    asset_schema = schema["components"]["schemas"]["AssetValuationResponse"]
    assert asset_schema["properties"]["final_economic_value"]["type"] == "string"


def _evidence_payload(**overrides):
    payload = {
        "scope_type": "account",
        "scope_key": "asset-1",
        "adjustment_type": "avaliacao_ativo",
        "method_component": "PLRA",
        "amount_impact": "20.00",
        "impact_base_value": "1000.00",
        "required_evidence_type": "documento_suporte",
        "evidence_status": "validada",
        "analyst_justification": "Valor suportado pelo documento.",
        "review_notes": None,
        "can_change_capag_status": False,
        "can_reverse_prudential_sign": False,
    }
    payload.update(overrides)
    return payload


def _asset_payload(*, evidence_id):
    return {
        "analysis_id": "analysis-1",
        "exercise_year": 2024,
        "account_code": "asset-1",
        "realizability_classification": (
            "liquidacao_forcada_exige_laudo"
        ),
        "valuation_required": True,
        "valuation_basis": "laudo_abnt_nbr_14653",
        "forced_liquidation_value": "450.00",
        "analyst_adjusted_value": None,
        "essentiality_status": "nao_essencial",
        "valuation_status": "validada",
        "evidence_id": evidence_id,
    }


def _client() -> TestClient:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        company = CompanyModel(
            id="company-1",
            legal_name="Empresa de teste",
            tax_id="00000000000100",
        )
        ecd_file = EcdFileModel(
            id="ecd-1",
            company_id=company.id,
            original_filename="teste.ecd",
            content_hash="sha256:evidence-api",
            layout="ECD_9",
            period_start=date(2024, 1, 1),
            period_end=date(2024, 12, 31),
        )
        analysis = AnalysisModel(
            id="analysis-1",
            company_id=company.id,
            ecd_file_id=ecd_file.id,
            methodology_version_id="metodologia-2024.1",
            status=ProcessingStatus.NOT_RUN.value,
        )
        exercise = ExerciseModel(
            analysis_id=analysis.id,
            year=2024,
            status=ProcessingStatus.NOT_RUN.value,
            methodology_version_id="metodologia-2024.1",
        )
        session.add_all([company, ecd_file, analysis, exercise])
        session.flush()
        session.add_all(
            [
                EcdI050AccountModel(
                    exercise_id=exercise.id,
                    account_code="asset-1",
                    account_name="Maquinas e equipamentos",
                    account_type="A",
                    account_nature="01",
                    level=5,
                    parent_account_code="asset-parent",
                    line_number=2,
                    source_line="|I050|...|",
                ),
                EcdI051ReferenceLinkModel(
                    exercise_id=exercise.id,
                    account_code="asset-1",
                    reference_code="1.02.03.01.06",
                    line_number=3,
                    source_line="|I051|1.02.03.01.06|",
                ),
                EcdI155BalanceModel(
                    exercise_id=exercise.id,
                    account_code="asset-1",
                    initial_balance=Decimal("800.00"),
                    initial_balance_indicator="D",
                    debit_amount=Decimal("0.00"),
                    credit_amount=Decimal("0.00"),
                    final_balance=Decimal("800.00"),
                    final_balance_indicator="D",
                    line_number=2,
                    source_line="|I155|...|",
                ),
                EcdI155BalanceModel(
                    exercise_id=exercise.id,
                    account_code="asset-1",
                    initial_balance=Decimal("1000.00"),
                    initial_balance_indicator="D",
                    debit_amount=Decimal("0.00"),
                    credit_amount=Decimal("0.00"),
                    final_balance=Decimal("1000.00"),
                    final_balance_indicator="D",
                    line_number=4,
                    source_line="|I155|...|",
                ),
            ]
        )
        session.commit()

    app = create_app()

    def override_session():
        with Session(engine) as session:
            yield session

    app.dependency_overrides[get_evidence_session] = override_session
    return TestClient(app)
