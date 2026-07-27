from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.domain import (
    CapagEAssessment,
    CapagEMethod,
    CapagEStatus,
    ComponentStatus,
)
from app.engine import calculate_capag_e_assessment
from app.export import (
    build_capag_assessment_workbook,
    serialize_capag_assessment_workbook,
)


def test_capag_excel_exports_partial_status_limitations_and_blocking_issues() -> None:
    assessment = calculate_capag_e_assessment(
        exercise_year=2024,
        method=CapagEMethod.FCA_PLRA,
        plra_value=Decimal("500000.00"),
        plra_status=ComponentStatus.CALCULATED,
        fco_value=Decimal("40000.00"),
        warnings=("Valor sujeito a revisao.",),
        blocking_issues=("EVIDENCIA_PENDENTE",),
        methodology_version_id="metodologia-2024.1",
    )

    workbook = build_capag_assessment_workbook(assessment)
    sheet = workbook["contrato_capag_e"]
    values = dict(zip((cell.value for cell in sheet[1]), (cell.value for cell in sheet[2])))

    assert workbook.sheetnames == ["contrato_capag_e"]
    assert values["metodo"] == "fca_plra"
    assert values["fca"] == "40000.00"
    assert values["status_fca"] == "parcial"
    assert values["capag_e"] == "540000.00"
    assert values["status_final"] == "parcial"
    assert "FCA parcial" in values["limitacoes_metodologicas"]
    assert values["warnings"] == "Valor sujeito a revisao."
    assert values["bloqueios"] == "EVIDENCIA_PENDENTE"
    assert values["sem_recalculo"] == "true"
    assert _has_no_formulas(workbook)


def test_capag_excel_serializes_snapshot_value_without_recalculation() -> None:
    assessment = CapagEAssessment(
        exercise_year=2024,
        method=CapagEMethod.FCA_PLRA,
        plra_value=Decimal("500000.00"),
        plra_status=ComponentStatus.CALCULATED,
        fca_value=Decimal("120000.00"),
        fca_status=ComponentStatus.CALCULATED,
        roa_value=None,
        roa_status=ComponentStatus.NOT_CALCULATED,
        capag_e_value=Decimal("111111.11"),
        capag_e_status=CapagEStatus.CALCULATED,
        unavailable_reason=None,
        calculation_basis="Snapshot externo controlado.",
        methodology_formula="CAPAG-E = PLRA + FCA",
        warnings=(),
        limitations=(),
        blocking_issues=(),
        methodology_version_id="metodologia-2024.1",
    )

    workbook = build_capag_assessment_workbook(assessment)

    assert workbook["contrato_capag_e"]["J2"].value == "111111.11"
    assert workbook["contrato_capag_e"]["J2"].value != "620000.00"
    assert _has_no_formulas(workbook)


def test_capag_excel_returns_readable_xlsx_bytes() -> None:
    assessment = calculate_capag_e_assessment(
        exercise_year=2024,
        method=CapagEMethod.ROA_PLRA,
        plra_value=Decimal("500000.00"),
        plra_status=ComponentStatus.CALCULATED,
        roa_value=Decimal("80000.00"),
        roa_status=ComponentStatus.CALCULATED,
        methodology_version_id="metodologia-2024.1",
    )

    payload = serialize_capag_assessment_workbook(assessment)
    workbook = load_workbook(BytesIO(payload))

    assert workbook["contrato_capag_e"]["J2"].value == "580000.00"
    assert workbook["contrato_capag_e"]["K2"].value == "calculado"
    assert _has_no_formulas(workbook)


def _has_no_formulas(workbook) -> bool:
    return all(
        not (isinstance(cell.value, str) and cell.value.startswith("="))
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )
