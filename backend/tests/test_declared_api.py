from decimal import Decimal
from datetime import date
from hashlib import sha256
from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, func, select
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.api.declared import get_declared_run_session, get_declared_snapshot_reader
from app.application import EcdImportIdentifiers, persist_parsed_ecd
from app.application.declared_service import (
    DeclaredAccountSnapshotView,
    DeclaredLayerSummary,
    DeclaredSnapshotsNotFound,
)
from app.domain import (
    BalanceComponent,
    BalanceLineStatus,
    BalanceRowStructuralStatus,
    DeclaredBalance,
    DeclaredBalanceRow,
    DeclaredBalanceStatus,
)
from app.main import app
from app.io import parse_ecd_file
from app.repositories import Base, DeclaredAccountSnapshot


BALANCE_FIXTURE = (
    Path(__file__).parent / "fixtures" / "ecd" / "balance_declared_valid.ecd"
)


class FakeDeclaredSnapshotReader:
    def get_summary(self, *, analysis_id: str, year: int) -> DeclaredLayerSummary:
        if analysis_id == "missing":
            raise DeclaredSnapshotsNotFound("Declared snapshot not found.")

        return DeclaredLayerSummary(
            analysis_id=analysis_id,
            year=year,
            total_accounts=2,
            status_counts={
                "MAPEADO": 1,
                "NAO_MAPEADO_METODOLOGICAMENTE": 1,
            },
            methodology_version_id="test-version",
        )

    def list_accounts(
        self,
        *,
        analysis_id: str,
        year: int,
    ) -> list[DeclaredAccountSnapshotView]:
        if analysis_id == "missing":
            raise DeclaredSnapshotsNotFound("Declared snapshot not found.")

        return [
            DeclaredAccountSnapshotView(
                account_code="1725",
                account_name="EMPRESTIMO - SICOOB CREDICITRUS - C",
                account_type="A",
                account_nature="02",
                account_level=5,
                parent_account_code="2.01.01.07",
                account_order=20,
                declared_reference_code="2.01.01.07.01",
                official_description="Emprestimos e financiamentos",
                official_reference_status="ATIVA",
                methodology_rule_applied="2.01.01.07.01",
                methodology_rule_status="ATIVA",
                purpose="FCO",
                treatment="tratamento_financiamento",
                base_value=Decimal("100000.00"),
                considered_value=Decimal("0.00"),
                final_status="MAPEADO",
                observation=None,
                recommended_action=None,
                methodology_version_id="test-version",
            ),
            DeclaredAccountSnapshotView(
                account_code="9999",
                account_name="Conta sem metodologia exata",
                account_type="A",
                account_nature="01",
                account_level=3,
                parent_account_code="9.99",
                account_order=30,
                declared_reference_code="9.99.99",
                official_description="Conta oficial de teste",
                official_reference_status="ATIVA",
                methodology_rule_applied=None,
                methodology_rule_status=None,
                purpose="FCO",
                treatment=None,
                base_value=Decimal("50.10"),
                considered_value=Decimal("50.10"),
                final_status="NAO_MAPEADO_METODOLOGICAMENTE",
                observation="Nenhuma regra metodologica exata encontrada.",
                recommended_action="revisar_metodologia",
                methodology_version_id="test-version",
            ),
        ]

def test_declared_accounts_endpoint_serializes_decimal_values_as_strings() -> None:
    app.dependency_overrides[get_declared_snapshot_reader] = FakeDeclaredSnapshotReader
    client = TestClient(app)

    response = client.get("/api/v1/analyses/analysis-1/exercises/2024/declared/accounts")

    app.dependency_overrides.clear()
    assert response.status_code == 200
    payload = response.json()
    assert payload["analysis_id"] == "analysis-1"
    assert payload["year"] == 2024
    assert payload["accounts"][0]["base_value"] == "100000.00"
    assert payload["accounts"][0]["considered_value"] == "0.00"
    assert payload["accounts"][0]["account_type"] == "A"
    assert payload["accounts"][0]["account_nature"] == "02"
    assert payload["accounts"][0]["account_level"] == 5
    assert payload["accounts"][0]["parent_account_code"] == "2.01.01.07"
    assert payload["accounts"][0]["account_order"] == 20
    assert payload["accounts"][0]["final_status"] == "MAPEADO"
    assert payload["accounts"][1]["final_status"] == "NAO_MAPEADO_METODOLOGICAMENTE"
    assert payload["accounts"][1]["recommended_action"] == "revisar_metodologia"


def test_declared_balance_endpoint_uses_specific_contract_and_decimal_strings(
    monkeypatch,
) -> None:
    monkeypatch.setattr("app.api.declared.get_declared_balance", lambda *args, **kwargs: _balance())
    client = TestClient(app)

    response = client.get(
        "/api/v1/analyses/analysis-1/exercises/2024/declared/balance/accounts"
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["balance_status"] == "VALIDO"
    assert payload["is_blocking"] is False
    assert payload["assets_final_amount"] == "800.00"
    assert payload["liabilities_and_equity_final_amount"] == "800.00"
    assert payload["difference"] == "0.00"
    assert payload["rows"][0]["aggregation_code"] == "ATIVO"
    assert payload["rows"][0]["children"][0]["reconciliation_status"] == "CONCILIADA"
    assert payload["rows"][0]["children"][0]["reconciled_amount"] == "800.00"
    assert "components" not in payload["rows"][0]["children"][0]


def test_declared_balance_components_endpoint_returns_audit_sources(monkeypatch) -> None:
    monkeypatch.setattr("app.api.declared.get_declared_balance", lambda *args, **kwargs: _balance())
    client = TestClient(app)

    response = client.get(
        "/api/v1/analyses/analysis-1/exercises/2024/declared/"
        "balance/accounts/AGL-CAIXA/components"
    )

    assert response.status_code == 200
    assert response.json() == {
        "analysis_id": "analysis-1",
        "year": 2024,
        "aggregation_code": "AGL-CAIXA",
        "rows": [
            {
                "account_code": "100",
                "account_name": "Caixa",
                "cost_center_code": "CC01",
                "final_amount": "800.00",
                "final_debit_credit_indicator": "D",
                "signed_final_amount": "800.00",
                "i052_line_number": 20,
                "i155_line_number": 25,
            }
        ],
    }


def test_declared_balance_components_endpoint_aggregates_totalizer_descendants(
    monkeypatch,
) -> None:
    monkeypatch.setattr("app.api.declared.get_declared_balance", lambda *args, **kwargs: _balance())
    client = TestClient(app)

    response = client.get(
        "/api/v1/analyses/analysis-1/exercises/2024/declared/"
        "balance/accounts/ATIVO/components"
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["aggregation_code"] == "ATIVO"
    assert payload["rows"] == [
        {
            "account_code": "100",
            "account_name": "Caixa",
            "cost_center_code": "CC01",
            "final_amount": "800.00",
            "final_debit_credit_indicator": "D",
            "signed_final_amount": "800.00",
            "i052_line_number": 20,
            "i155_line_number": 25,
        }
    ]


def test_declared_balance_components_endpoint_returns_explicit_missing_row(monkeypatch) -> None:
    monkeypatch.setattr("app.api.declared.get_declared_balance", lambda *args, **kwargs: _balance())
    client = TestClient(app)

    response = client.get(
        "/api/v1/analyses/analysis-1/exercises/2024/declared/"
        "balance/accounts/INEXISTENTE/components"
    )

    assert response.status_code == 404
    assert response.json()["detail"]["error_code"] == "DECLARED_BALANCE_ROW_NOT_FOUND"


def test_declared_balance_get_reads_database_without_writing_snapshot() -> None:
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    raw = BALANCE_FIXTURE.read_bytes()
    with Session(engine) as session:
        persist_parsed_ecd(
            session,
            parsed_ecd=parse_ecd_file(BALANCE_FIXTURE),
            identifiers=EcdImportIdentifiers(
                company_id="company-api-balance",
                ecd_file_id="ecd-api-balance",
                analysis_id="analysis-api-balance",
                methodology_version_id="metodologia-2024.1",
                original_filename=BALANCE_FIXTURE.name,
                content_hash=f"sha256:{sha256(raw).hexdigest()}",
            ),
            original_content=raw,
            parser_version="2.0.0",
        )

    def session_override():
        with Session(engine) as session:
            yield session

    app.dependency_overrides[get_declared_run_session] = session_override
    try:
        response = TestClient(app).get(
            "/api/v1/analyses/analysis-api-balance/exercises/2024/"
            "declared/balance/accounts"
        )
    finally:
        app.dependency_overrides.clear()

    with Session(engine) as session:
        snapshot_count = session.scalar(
            select(func.count()).select_from(DeclaredAccountSnapshot)
        )

    assert response.status_code == 200
    assert response.json()["balance_status"] == "VALIDO"
    assert snapshot_count == 0


def test_declared_summary_endpoint_returns_status_counts() -> None:
    app.dependency_overrides[get_declared_snapshot_reader] = FakeDeclaredSnapshotReader
    client = TestClient(app)

    response = client.get("/api/v1/analyses/analysis-1/exercises/2024/declared")

    app.dependency_overrides.clear()
    assert response.status_code == 200
    assert response.json() == {
        "analysis_id": "analysis-1",
        "year": 2024,
        "total_accounts": 2,
        "status_counts": {
            "MAPEADO": 1,
            "NAO_MAPEADO_METODOLOGICAMENTE": 1,
        },
        "methodology_version_id": "test-version",
    }


def test_declared_endpoint_maps_missing_snapshot_to_explicit_error() -> None:
    app.dependency_overrides[get_declared_snapshot_reader] = FakeDeclaredSnapshotReader
    client = TestClient(app)

    response = client.get("/api/v1/analyses/missing/exercises/2024/declared")

    app.dependency_overrides.clear()
    assert response.status_code == 404
    assert response.json()["detail"] == {
        "error_code": "DECLARED_SNAPSHOT_NOT_FOUND",
        "message": "Declared snapshot not found.",
    }


def test_declared_excel_endpoint_downloads_workbook_from_snapshots() -> None:
    app.dependency_overrides[get_declared_snapshot_reader] = FakeDeclaredSnapshotReader
    client = TestClient(app)

    response = client.get("/api/v1/analyses/analysis-1/exercises/2024/declared/export.xlsx")

    app.dependency_overrides.clear()
    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert (
        response.headers["content-disposition"]
        == 'attachment; filename="capag-declarada-analysis-1-2024.xlsx"'
    )

    workbook = load_workbook(BytesIO(response.content))
    assert workbook["resumo_executivo"]["B1"].value == "analysis-1"
    assert workbook["campos_resultado"]["A2"].value == "1725"
    assert workbook["campos_resultado"]["I2"].value == "100000.00"
    assert workbook["campos_resultado"]["J2"].value == "0.00"
    assert workbook["log_auditoria"]["B2"].value == "snapshot_declarado_persistido"
    assert _has_no_formulas(workbook)


def test_declared_openapi_contains_declared_contracts() -> None:
    paths = app.openapi()["paths"]

    assert "/api/v1/analyses/{analysis_id}/exercises/{year}/declared" in paths
    assert "/api/v1/analyses/{analysis_id}/exercises/{year}/declared/accounts" in paths
    assert "/api/v1/analyses/{analysis_id}/exercises/{year}/declared/balance/accounts" in paths
    assert (
        "/api/v1/analyses/{analysis_id}/exercises/{year}/declared/"
        "balance/accounts/{aggregation_code}/components"
    ) in paths
    assert "/api/v1/analyses/{analysis_id}/exercises/{year}/declared/export.xlsx" in paths
    accounts_operation = paths[
        "/api/v1/analyses/{analysis_id}/exercises/{year}/declared/accounts"
    ]["get"]
    assert accounts_operation["responses"]["200"]["description"] == "Successful Response"
    balance_operation = paths[
        "/api/v1/analyses/{analysis_id}/exercises/{year}/declared/balance/accounts"
    ]["get"]
    assert balance_operation["responses"]["200"]["content"]["application/json"]["schema"] == {
        "$ref": "#/components/schemas/DeclaredBalanceResponse"
    }


def _balance() -> DeclaredBalance:
    component = BalanceComponent(
        account_code="100",
        account_name="Caixa",
        cost_center_code="CC01",
        final_amount=Decimal("800.00"),
        final_debit_credit_indicator="D",
        signed_final_amount=Decimal("800.00"),
        i052_line_number=20,
        i155_line_number=25,
    )
    detail = DeclaredBalanceRow(
        aggregation_code="AGL-CAIXA",
        aggregation_code_type="D",
        aggregation_level=2,
        parent_aggregation_code="ATIVO",
        balance_group="A",
        description="Caixa",
        initial_amount=Decimal("100.00"),
        initial_debit_credit_indicator="D",
        signed_initial_amount=Decimal("100.00"),
        final_amount=Decimal("800.00"),
        final_debit_credit_indicator="D",
        signed_final_amount=Decimal("800.00"),
        explanatory_note_reference=None,
        line_number=31,
        structural_status=BalanceRowStructuralStatus.VALIDA,
        reconciliation_status=BalanceLineStatus.CONCILIADA,
        reconciled_amount=Decimal("800.00"),
        difference=Decimal("0.00"),
        component_count=1,
        components=(component,),
        children=(),
    )
    root = DeclaredBalanceRow(
        aggregation_code="ATIVO",
        aggregation_code_type="T",
        aggregation_level=1,
        parent_aggregation_code=None,
        balance_group="A",
        description="Ativo",
        initial_amount=Decimal("100.00"),
        initial_debit_credit_indicator="D",
        signed_initial_amount=Decimal("100.00"),
        final_amount=Decimal("800.00"),
        final_debit_credit_indicator="D",
        signed_final_amount=Decimal("800.00"),
        explanatory_note_reference=None,
        line_number=30,
        structural_status=BalanceRowStructuralStatus.VALIDA,
        reconciliation_status=None,
        reconciled_amount=None,
        difference=None,
        component_count=0,
        components=(),
        children=(detail,),
    )
    return DeclaredBalance(
        year=2024,
        status=DeclaredBalanceStatus.VALIDO,
        is_blocking=False,
        j005_period_start=date(2024, 1, 1),
        j005_period_end=date(2024, 12, 31),
        assets_final_amount=Decimal("800.00"),
        liabilities_and_equity_final_amount=Decimal("800.00"),
        difference=Decimal("0.00"),
        rows=(root,),
        limitations=(),
    )


def _has_no_formulas(workbook) -> bool:
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    return False

    return True
