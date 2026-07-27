from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.domain import (
    ComponentStatus,
    RoaAuditRow,
    RoaBlock,
    RoaCalculation,
    RoaComponentSummary,
    RoaPendingGroup,
    RoaRowStatus,
)
from app.engine import calculate_capag_e_assessment
from app.export import build_roa_workbook, serialize_roa_workbook


def test_roa_excel_exports_summary_audit_pressures_pending_and_capag() -> None:
    calculation = _calculation()
    workbook = build_roa_workbook(
        calculation,
        analysis_id="analysis-1",
        capag_assessment=_assessment(),
    )
    summary_rows = _rows_by_type(workbook["roa_resumo"])
    audit = _row_dict(workbook["roa_auditoria"])
    pressure = _row_dict(workbook["roa_pressoes_caixa"])

    assert workbook.sheetnames == [
        "roa_resumo",
        "roa_auditoria",
        "roa_pressoes_caixa",
    ]
    assert summary_rows["calculo"]["roa_final"] == "75.00"
    assert summary_rows["calculo"]["status_roa"] == "bloqueado_por_evidencia"
    assert summary_rows["calculo"]["plra"] == "500.00"
    assert summary_rows["calculo"]["capag_e"] is None
    assert summary_rows["calculo"]["status_capag_e"] == "bloqueado"
    assert "J150" in summary_rows["calculo"]["limitacoes"]
    assert summary_rows["componente"]["valor_componente"] == "100.00"
    assert summary_rows["pendencia"]["codigo_pendencia"] == "EVIDENCIA_PENDENTE"
    assert summary_rows["pendencia"]["bloqueia_roa"] == "true"
    assert audit["efeito_roa"] == "100.00"
    assert pressure["valor_pressao"] == "25.00"
    assert pressure["efeito_roa"] == "-25.00"
    assert _has_no_formulas(workbook)


def test_roa_excel_serializes_snapshot_values_without_recalculation() -> None:
    payload = serialize_roa_workbook(
        _calculation(),
        analysis_id="analysis-1",
        capag_assessment=_assessment(),
    )
    workbook = load_workbook(BytesIO(payload))
    summary = _rows_by_type(workbook["roa_resumo"])["calculo"]

    assert summary["roa_preliminar"] == "100.00"
    assert summary["pressoes_caixa"] == "25.00"
    assert summary["roa_final"] == "75.00"
    assert _has_no_formulas(workbook)


def _calculation() -> RoaCalculation:
    return RoaCalculation(
        exercise_year=2024,
        gross_revenue=Decimal("100"),
        deductions=Decimal("0"),
        revenue_taxes=Decimal("0"),
        net_operating_revenue=Decimal("100"),
        operating_costs=Decimal("0"),
        operating_expenses=Decimal("0"),
        financial_result=Decimal("0"),
        non_operating_result=Decimal("0"),
        cash_pressure_adjustments=Decimal("25"),
        roa_preliminary=Decimal("100"),
        roa_final=Decimal("75"),
        status=ComponentStatus.BLOCKED_BY_EVIDENCE,
        component_summaries=(
            RoaComponentSummary(
                block=RoaBlock.GROSS_REVENUE,
                component_code="receita_vendas_servicos",
                component_label="Receita de vendas e serviços",
                value=Decimal("100"),
                account_count=1,
            ),
        ),
        audit_rows=(
            RoaAuditRow(
                account_code="sales",
                account_name="Receita",
                reference_code="3.01.01.01.01.04",
                reference_description="Receita de vendas",
                roa_block=RoaBlock.GROSS_REVENUE,
                component_roa="receita_vendas_servicos",
                component_label="Receita de vendas e serviços",
                base_value=Decimal("100"),
                signed_value=Decimal("100"),
                treatment="incluir_automaticamente",
                final_status=RoaRowStatus.INCLUDED,
                pending_reason=None,
                evidence_id=None,
                line_reference=10,
                macrogroup="RECEITA_OPERACIONAL",
                required_evidence_type="documento_fiscal_receita",
                source_detail="Natureza de saldo I155: C.",
            ),
            RoaAuditRow(
                account_code="pressure-1",
                account_name="Fornecedores vencidos",
                reference_code=None,
                reference_description=None,
                roa_block=RoaBlock.CASH_PRESSURES,
                component_roa="fornecedores_vencidos",
                component_label="Fornecedores vencidos",
                base_value=Decimal("25"),
                signed_value=Decimal("-25"),
                treatment="incluir_automaticamente",
                final_status=RoaRowStatus.PENDING_EVIDENCE,
                pending_reason="evidencia_roa_material_ausente",
                evidence_id=None,
                line_reference=20,
                macrogroup="PRESSAO_COMPLEMENTAR_CAIXA",
                required_evidence_type="relatorio_fornecedores_vencidos",
                source_detail="Cadastro controlado.",
            ),
        ),
        pending_groups=(
            RoaPendingGroup(
                code="EVIDENCIA_PENDENTE",
                message="Pressão sem evidência.",
                account_code="pressure-1",
                reference_code=None,
                blocks_roa=True,
                materiality_level="alta",
                evidence_id=None,
            ),
        ),
        alerts=("Pressão material.",),
        limitations=(
            "Conferencia J150 indisponivel; ROA calculado a partir de I155.",
        ),
        methodology_version_id="metodologia-2024.1",
    )


def _assessment():
    return calculate_capag_e_assessment(
        exercise_year=2024,
        method="roa_plra",
        plra_value=Decimal("500"),
        plra_status=ComponentStatus.CALCULATED,
        roa_value=Decimal("75"),
        roa_status=ComponentStatus.BLOCKED_BY_EVIDENCE,
        methodology_version_id="metodologia-2024.1",
    )


def _row_dict(sheet, row_number: int = 2) -> dict[str, object]:
    headers = [cell.value for cell in sheet[1]]
    return dict(zip(headers, (cell.value for cell in sheet[row_number])))


def _rows_by_type(sheet) -> dict[str, dict[str, object]]:
    return {
        str(row["tipo_registro"]): row
        for row_number in range(2, sheet.max_row + 1)
        if (row := _row_dict(sheet, row_number))["tipo_registro"] is not None
    }


def _has_no_formulas(workbook) -> bool:
    return not any(
        isinstance(cell.value, str) and cell.value.startswith("=")
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )
