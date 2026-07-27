from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.domain import (
    CashFlowDirection,
    ComponentStatus,
    DfcActivity,
    DfcAuditRow,
    DfcCalculation,
    DfcComponentSummary,
    DfcPendingIssue,
    DfcRowStatus,
)
from app.export import build_dfc_workbook, serialize_dfc_workbook


def test_dfc_excel_exports_summary_components_pending_issues_and_audit() -> None:
    workbook = build_dfc_workbook(_calculation(), analysis_id="analysis-1")
    summary = workbook["dfc_resumo"]
    audit = workbook["dfc_auditoria"]
    summary_rows = _rows_by_type(summary)
    audit_row = _row_dict(audit)

    assert workbook.sheetnames == ["dfc_resumo", "dfc_auditoria"]
    assert summary_rows["calculo"]["fca"] == "75.00"
    assert summary_rows["calculo"]["status_fca"] == "bloqueado_por_pendencia"
    assert summary_rows["calculo"]["sem_recalculo"] == "true"
    assert summary_rows["componente"]["codigo_componente"] == "recebimentos_clientes"
    assert summary_rows["componente"]["valor_componente"] == "100.00"
    assert summary_rows["pendencia"]["codigo_pendencia"] == "regra_ausente"
    assert summary_rows["pendencia"]["bloqueia_fca"] == "true"
    assert audit_row["valor_movimento"] == "100.00"
    assert audit_row["valor_incluido"] == "100.00"
    assert audit_row["status_fca"] == "bloqueado_por_pendencia"
    assert _has_no_formulas(workbook)


def test_dfc_excel_serializes_readable_snapshot_without_recalculation() -> None:
    payload = serialize_dfc_workbook(_calculation(), analysis_id="analysis-1")
    workbook = load_workbook(BytesIO(payload))

    summary = _rows_by_type(workbook["dfc_resumo"])["calculo"]
    assert summary["fluxo_operacional"] == "100.00"
    assert summary["fluxo_investimento"] == "-25.00"
    assert summary["fca"] == "75.00"
    assert _has_no_formulas(workbook)


def _calculation() -> DfcCalculation:
    return DfcCalculation(
        exercise_year=2024,
        automatic_value=Decimal("75.00"),
        operational_flow=Decimal("100.00"),
        investment_flow=Decimal("-25.00"),
        financing_flow=Decimal("0.00"),
        manual_adjustments_value=Decimal("0.00"),
        fca_value=Decimal("75.00"),
        status=ComponentStatus.BLOCKED_BY_PENDING,
        component_summaries=(
            DfcComponentSummary(
                activity=DfcActivity.OPERATIONAL,
                component_code="recebimentos_clientes",
                component_label="Recebimentos de clientes",
                value=Decimal("100.00"),
                movement_count=1,
            ),
        ),
        audit_rows=(
            DfcAuditRow(
                entry_number="LCTO-1",
                entry_date=date(2024, 1, 31),
                cash_account_code="cash",
                cash_flow_direction=CashFlowDirection.INFLOW,
                counterparty_account_code="clients",
                counterparty_account_name="Clientes",
                counterparty_reference_code="3.01.01.01.01.04",
                dfc_activity=DfcActivity.OPERATIONAL,
                dfc_component_code="recebimentos_clientes",
                dfc_component_label="Recebimentos de clientes",
                movement_value=Decimal("100.00"),
                included_value=Decimal("100.00"),
                final_status=DfcRowStatus.INCLUDED,
                pending_reason=None,
                history="Recebimento",
                line_number=10,
            ),
        ),
        pending_issues=(
            DfcPendingIssue(
                code="regra_ausente",
                message="Movimento material sem regra.",
                entry_number="LCTO-2",
                line_number=11,
                materiality_level="alta",
                blocks_fca=True,
            ),
        ),
        alerts=("Fluxo revisado.",),
        limitations=("FCA bloqueado até decisão.",),
        methodology_version_id="metodologia-2024.1",
    )


def _row_dict(sheet, row_number: int = 2) -> dict[str, object]:
    headers = [cell.value for cell in sheet[1]]
    return dict(zip(headers, (cell.value for cell in sheet[row_number])))


def _rows_by_type(sheet) -> dict[str, dict[str, object]]:
    return {
        str(row["tipo_registro"]): row
        for row_number in range(2, sheet.max_row + 1)
        if (row := _row_dict(sheet, row_number))["tipo_registro"] is not None
    }


def _has_no_formulas(workbook) -> bool:
    return not any(
        isinstance(cell.value, str) and cell.value.startswith("=")
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )
