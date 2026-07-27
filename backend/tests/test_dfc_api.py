from datetime import date
from decimal import Decimal
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.api.dfc import get_dfc_session
from app.domain import ProcessingStatus
from app.main import create_app
from app.repositories import (
    AnalysisModel,
    Base,
    CompanyModel,
    EcdFileModel,
    EcdI050AccountModel,
    EcdI051ReferenceLinkModel,
    EcdI200EntryModel,
    EcdI250EntryItemModel,
    ExerciseModel,
)


BASE_PATH = "/api/v1/analyses/analysis-1/exercises/2024/dfc"


def test_dfc_api_runs_exposes_audit_and_applies_manual_decision() -> None:
    client = _client()

    run_response = client.post(f"{BASE_PATH}/run")

    assert run_response.status_code == 200
    first = run_response.json()
    assert first["fca_value"] == "0.00"
    assert first["fca_status"] == "bloqueado_por_pendencia"
    assert first["audit_rows"][0]["movement_value"] == "100.00"
    assert first["audit_rows"][0]["final_status"] == "nao_classificado"

    decision_response = client.post(
        f"{BASE_PATH}/decisions",
        json={
            "action": "incluir",
            "entry_number": "LCTO-1",
            "line_number": 6,
            "activity": "operacional",
            "component_code": "recebimentos_clientes",
            "justification": "Recebimento conciliado com extrato.",
        },
    )

    assert decision_response.status_code == 200
    decided = decision_response.json()
    assert decided["fca_value"] == "100.00"
    assert decided["fca_status"] == "calculado"
    assert decided["audit_rows"][0]["final_status"] == "decisao_manual_aplicada"
    assert client.get(BASE_PATH).json() == decided


def test_dfc_api_rejects_invalid_decision_contract() -> None:
    client = _client()
    client.post(f"{BASE_PATH}/run")

    response = client.post(
        f"{BASE_PATH}/decisions",
        json={
            "action": "incluir",
            "entry_number": "LCTO-1",
            "line_number": 6,
            "justification": "Sem classificacao.",
        },
    )

    assert response.status_code == 422


def test_dfc_api_returns_not_found_before_first_run() -> None:
    response = _client().get(BASE_PATH)

    assert response.status_code == 404
    assert response.json()["detail"]["error_code"] == "DFC_NOT_FOUND"


def test_dfc_api_exports_persisted_snapshot_without_formulas() -> None:
    client = _client()
    client.post(f"{BASE_PATH}/run")

    response = client.get(f"{BASE_PATH}/export.xlsx")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["dfc_resumo", "dfc_auditoria"]
    assert workbook["dfc_resumo"]["A2"].value == "calculo"
    assert workbook["dfc_resumo"]["D2"].value == "0.00"
    assert workbook["dfc_resumo"]["E2"].value == "bloqueado_por_pendencia"
    assert workbook["dfc_auditoria"]["M2"].value == "100.00"
    assert workbook["dfc_auditoria"]["N2"].value == "0.00"


def test_dfc_openapi_exposes_governed_paths_and_decimal_strings() -> None:
    schema = create_app().openapi()
    path = BASE_PATH.replace("analysis-1", "{analysis_id}").replace(
        "2024", "{year}"
    )

    assert "get" in schema["paths"][path]
    assert "post" in schema["paths"][f"{path}/run"]
    assert "post" in schema["paths"][f"{path}/decisions"]
    assert "get" in schema["paths"][f"{path}/export.xlsx"]
    response_schema = schema["components"]["schemas"]["DfcCalculationResponse"]
    assert response_schema["properties"]["fca_value"]["type"] == "string"
    row_schema = schema["components"]["schemas"]["DfcAuditRowResponse"]
    assert row_schema["properties"]["included_value"]["type"] == "string"


def _client() -> TestClient:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        company = CompanyModel(
            id="company-1",
            legal_name="Empresa de teste",
            tax_id="00000000000100",
        )
        ecd_file = EcdFileModel(
            id="ecd-1",
            company_id=company.id,
            original_filename="teste.ecd",
            content_hash="sha256:dfc-api",
            layout="ECD_9",
            period_start=date(2024, 1, 1),
            period_end=date(2024, 12, 31),
        )
        analysis = AnalysisModel(
            id="analysis-1",
            company_id=company.id,
            ecd_file_id=ecd_file.id,
            methodology_version_id="metodologia-2024.1",
            status=ProcessingStatus.NOT_RUN.value,
        )
        exercise = ExerciseModel(
            analysis_id=analysis.id,
            year=2024,
            status=ProcessingStatus.NOT_RUN.value,
            methodology_version_id="metodologia-2024.1",
        )
        session.add_all([company, ecd_file, analysis, exercise])
        session.flush()
        session.add_all(
            [
                EcdI050AccountModel(
                    exercise_id=exercise.id,
                    account_code="cash",
                    account_name="Banco",
                    account_type="A",
                    account_nature="01",
                    level=5,
                    parent_account_code=None,
                    line_number=2,
                    source_line="|I050|cash|",
                ),
                EcdI051ReferenceLinkModel(
                    exercise_id=exercise.id,
                    account_code="cash",
                    reference_code="1.01.01.02.01",
                    line_number=3,
                    source_line="|I051|1.01.01.02.01|",
                ),
                EcdI050AccountModel(
                    exercise_id=exercise.id,
                    account_code="other",
                    account_name="Tributo a recuperar",
                    account_type="A",
                    account_nature="01",
                    level=5,
                    parent_account_code=None,
                    line_number=4,
                    source_line="|I050|other|",
                ),
                EcdI051ReferenceLinkModel(
                    exercise_id=exercise.id,
                    account_code="other",
                    reference_code="1.01.02.03.02",
                    line_number=5,
                    source_line="|I051|1.01.02.03.02|",
                ),
            ]
        )
        entry = EcdI200EntryModel(
            exercise_id=exercise.id,
            entry_number="LCTO-1",
            entry_date=date(2024, 1, 31),
            total_amount=Decimal("100.00"),
            line_number=5,
            source_line="|I200|LCTO-1|",
        )
        session.add(entry)
        session.flush()
        session.add_all(
            [
                EcdI250EntryItemModel(
                    entry_id=entry.id,
                    account_code="cash",
                    counterparty_account_code=None,
                    amount=Decimal("100.00"),
                    debit_credit_indicator="D",
                    history="Recebimento",
                    line_number=5,
                    source_line="|I250|cash|",
                ),
                EcdI250EntryItemModel(
                    entry_id=entry.id,
                    account_code="other",
                    counterparty_account_code=None,
                    amount=Decimal("100.00"),
                    debit_credit_indicator="C",
                    history="Recebimento",
                    line_number=6,
                    source_line="|I250|other|",
                ),
            ]
        )
        session.commit()

    app = create_app()

    def override_session():
        with Session(engine) as session:
            yield session

    app.dependency_overrides[get_dfc_session] = override_session
    return TestClient(app)
