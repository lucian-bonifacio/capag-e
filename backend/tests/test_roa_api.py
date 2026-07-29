from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.api.roa import get_roa_session
from app.domain import (
    ComponentStatus,
    DfcCalculation,
    PlraCalculation,
    ProcessingStatus,
)
from app.main import create_app
from app.repositories import (
    AnalysisModel,
    Base,
    CompanyModel,
    EcdFileModel,
    EcdI050AccountModel,
    EcdI051ReferenceLinkModel,
    EcdI155BalanceModel,
    ExerciseModel,
    add_dfc_calculation,
    add_plra_calculation,
)


BASE_PATH = "/api/v1/analyses/analysis-1/exercises/2024/roa"


def test_roa_api_runs_persists_and_applies_conditional_decision() -> None:
    client = _client()

    run_response = client.post(f"{BASE_PATH}/run")

    assert run_response.status_code == 200
    first = run_response.json()
    assert first["gross_revenue"] == "100.00"
    assert first["non_operating_result"] == "0.00"
    assert first["roa_final"] == "100.00"
    assert first["roa_status"] == "bloqueado_por_pendencia"
    assert first["audit_rows"][1]["final_status"] == "pendente_revisao"
    assert first["capag_assessment"]["roa_value"] == "100.00"
    assert first["capag_assessment"]["capag_e_value"] is None

    decision_response = client.post(
        f"{BASE_PATH}/decisions",
        json={
            "action": "incluir",
            "account_code": "other-revenue",
            "justification": "Receita não recorrente conciliada.",
        },
    )

    assert decision_response.status_code == 200
    decided = decision_response.json()
    assert decided["non_operating_result"] == "20.00"
    assert decided["roa_final"] == "120.00"
    assert decided["roa_status"] == "calculado"
    assert decided["audit_rows"][1]["final_status"] == "decisao_manual_aplicada"
    assert decided["capag_assessment"]["method"] == "roa_plra"
    assert decided["capag_assessment"]["capag_e_value"] == "620.00"
    assert decided["capag_assessment"]["capag_e_status"] == "calculado"
    assert client.get(BASE_PATH).json() == decided


def test_roa_api_rejects_manual_inclusion_without_methodology_rule() -> None:
    client = _client(include_no_rule=True)
    client.post(f"{BASE_PATH}/run")

    response = client.post(
        f"{BASE_PATH}/decisions",
        json={
            "action": "incluir",
            "account_code": "no-rule",
            "justification": "Classificação ainda não governada.",
        },
    )

    assert response.status_code == 422
    assert (
        response.json()["detail"]["error_code"]
        == "ROA_DECISION_ERROR"
    )


def test_roa_api_preserves_fca_comparison_when_snapshot_exists() -> None:
    response = _client(include_fca=True).post(f"{BASE_PATH}/run")

    assert response.status_code == 200
    assessment = response.json()["capag_assessment"]
    assert assessment["method"] == "comparativo_fca_roa"
    assert assessment["fca_value"] == "120.00"
    assert assessment["roa_value"] == "100.00"
    assert "PLRA+FCA=620.00" in assessment["calculation_basis"]
    assert "PLRA+ROA=indisponivel" in assessment["calculation_basis"]


def test_roa_api_returns_not_found_before_first_run() -> None:
    response = _client().get(BASE_PATH)

    assert response.status_code == 404
    assert response.json()["detail"]["error_code"] == "ROA_NOT_FOUND"


def test_roa_api_exports_persisted_snapshot_without_formulas() -> None:
    client = _client()
    client.post(f"{BASE_PATH}/run")

    response = client.get(f"{BASE_PATH}/export.xlsx")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == [
        "roa_resumo",
        "roa_auditoria",
        "roa_pressoes_caixa",
    ]
    assert workbook["roa_resumo"]["A2"].value == "calculo"
    assert workbook["roa_resumo"]["N2"].value == "100.00"
    assert workbook["roa_auditoria"]["K2"].value == "100.00"
    assert not any(
        isinstance(cell.value, str) and cell.value.startswith("=")
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )


def test_roa_openapi_exposes_governed_paths_and_decimal_strings() -> None:
    schema = create_app().openapi()
    path = BASE_PATH.replace("analysis-1", "{analysis_id}").replace(
        "2024",
        "{year}",
    )

    assert "get" in schema["paths"][path]
    assert "post" in schema["paths"][f"{path}/run"]
    assert "post" in schema["paths"][f"{path}/decisions"]
    assert "get" in schema["paths"][f"{path}/export.xlsx"]
    response_schema = schema["components"]["schemas"]["RoaCalculationResponse"]
    assert response_schema["properties"]["roa_final"]["type"] == "string"
    row_schema = schema["components"]["schemas"]["RoaAuditRowResponse"]
    assert row_schema["properties"]["signed_value"]["type"] == "string"


def _client(
    *,
    include_no_rule: bool = False,
    include_fca: bool = False,
) -> TestClient:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        company = CompanyModel(
            id="company-1",
            legal_name="Empresa de teste",
            tax_id="00000000000100",
        )
        ecd_file = EcdFileModel(
            id="ecd-1",
            company_id=company.id,
            original_filename="teste.ecd",
            content_hash=f"sha256:roa-api:{include_no_rule}:{include_fca}",
            layout="ECD_9",
            period_start=date(2024, 1, 1),
            period_end=date(2024, 12, 31),
        )
        analysis = AnalysisModel(
            id="analysis-1",
            company_id=company.id,
            ecd_file_id=ecd_file.id,
            methodology_version_id="metodologia-2024.1",
            status=ProcessingStatus.NOT_RUN.value,
        )
        exercise = ExerciseModel(
            analysis_id=analysis.id,
            year=2024,
            status=ProcessingStatus.NOT_RUN.value,
            methodology_version_id="metodologia-2024.1",
        )
        session.add_all([company, ecd_file, analysis, exercise])
        session.flush()
        _add_result_account(
            session,
            exercise_id=exercise.id,
            account_code="sales",
            account_name="Receita de vendas",
            reference_code="3.01.01.01.01.04",
            debit="0",
            credit="100",
            line_number=10,
        )
        _add_result_account(
            session,
            exercise_id=exercise.id,
            account_code="other-revenue",
            account_name="Outras receitas",
            reference_code="3.01.01.05.01.01",
            debit="0",
            credit="20",
            line_number=20,
        )
        if include_no_rule:
            _add_result_account(
                session,
                exercise_id=exercise.id,
                account_code="no-rule",
                account_name="Conta sem regra",
                reference_code="3.99.99",
                debit="10",
                credit="0",
                line_number=30,
            )
        add_plra_calculation(
            session,
            exercise_id=exercise.id,
            calculation=PlraCalculation(
                analysis_id=analysis.id,
                exercise_year=2024,
                gross_assets_value=Decimal("0"),
                gross_economic_liabilities_value=Decimal("0"),
                adjusted_assets_value=Decimal("0"),
                plr_gross_value=Decimal("0"),
                plra_value=Decimal("500"),
                plra_status=ComponentStatus.CALCULATED,
                calculation_formula="snapshot patrimonial",
                account_rows=(),
                pending_accounts=(),
                warnings=(),
                limitations=(),
                blocking_issues=(),
                balance_status="VALIDO",
                methodology_version_id="metodologia-2024.1",
                calculated_at=datetime(2026, 7, 24, tzinfo=timezone.utc),
            ),
        )
        if include_fca:
            add_dfc_calculation(
                session,
                exercise_id=exercise.id,
                analysis_id=analysis.id,
                calculation=DfcCalculation(
                    exercise_year=2024,
                    automatic_value=Decimal("120"),
                    operational_flow=Decimal("120"),
                    investment_flow=Decimal("0"),
                    financing_flow=Decimal("0"),
                    manual_adjustments_value=Decimal("0"),
                    fca_value=Decimal("120"),
                    status=ComponentStatus.CALCULATED,
                    component_summaries=(),
                    audit_rows=(),
                    pending_issues=(),
                    alerts=(),
                    limitations=(),
                    methodology_version_id="metodologia-2024.1",
                ),
            )
        session.commit()

    app = create_app()

    def override_session():
        with Session(engine) as session:
            yield session

    app.dependency_overrides[get_roa_session] = override_session
    return TestClient(app)


def _add_result_account(
    session: Session,
    *,
    exercise_id: int,
    account_code: str,
    account_name: str,
    reference_code: str,
    debit: str,
    credit: str,
    line_number: int,
) -> None:
    balance_nature = "C" if Decimal(credit) > Decimal(debit) else "D"
    session.add(
        EcdI050AccountModel(
            exercise_id=exercise_id,
            account_code=account_code,
            account_name=account_name,
            account_type="A",
            account_nature="04",
            level=5,
            parent_account_code=None,
            line_number=line_number,
            source_line=f"|I050|{account_code}|",
        )
    )
    session.add(
        EcdI051ReferenceLinkModel(
            exercise_id=exercise_id,
            account_code=account_code,
            reference_code=reference_code,
            line_number=line_number + 1,
            source_line=f"|I051|{reference_code}|",
        )
    )
    session.add(
        EcdI155BalanceModel(
            exercise_id=exercise_id,
            account_code=account_code,
            initial_balance=Decimal("0"),
            initial_balance_indicator=balance_nature,
            debit_amount=Decimal(debit),
            credit_amount=Decimal(credit),
            final_balance=Decimal("0"),
            final_balance_indicator=balance_nature,
            line_number=line_number + 2,
            source_line=f"|I155|{account_code}|",
        )
    )
