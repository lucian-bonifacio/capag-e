from __future__ import annotations

import argparse
from datetime import date, datetime
import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EXPECTED_SOURCE_HASH = (
    "0c66a19ce859cdc7a1eee137896243100cbaa26239ffa8ed3044762f3e359397"
)
OFFICIAL_VERSION_ID = "sped-ecf-11-ac2024-pj-geral-2025-11-09-v1"
METHODOLOGY_VERSION_ID = "metodologia-2024.1"
SOURCE_CALENDAR_YEAR = 2024
SOURCE_SHEETS = ("L100A", "L300A")
EXPECTED_SHEET_COUNTS = {"L100A": 722, "L300A": 387}
SOURCE_NATURES = {
    "1": "ATIVO",
    "2": "PASSIVO",
    "3": "PATRIMONIO_LIQUIDO",
    "4": "RESULTADO",
}
REQUIRED_FIELDS = [
    "reference_code",
    "official_description",
    "parent_reference_code",
    "level",
    "nature",
    "valid_from",
    "valid_to",
    "layout",
    "entity_type",
    "source",
    "status",
    "methodology_version_id",
    "source_sheet",
    "source_type",
    "source_nature_code",
    "source_valid_from",
    "source_valid_to",
    "official_guidance",
    "validation_notes",
]
SOURCE_URL = (
    "https://www.gov.br/sped/pt-br/assuntos/escrituracoes-digitais/ecf/"
    "manuais-e-documentos-tecnicos/"
    "tabelas_dinamicas_ecf_leiaute_11_09_11_2025_ac_2024_sit_esp_2025.xlsx/"
    "@@display-file/file"
)


class OfficialReferenceImportError(RuntimeError):
    pass


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    payload = build_payload(args.source)
    args.output.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def build_payload(source_path: Path) -> dict[str, Any]:
    source_hash = hashlib.sha256(source_path.read_bytes()).hexdigest()
    if source_hash != EXPECTED_SOURCE_HASH:
        raise OfficialReferenceImportError(
            f"Unexpected source SHA-256: {source_hash}."
        )

    workbook = load_workbook(source_path, read_only=True, data_only=True)
    records: list[dict[str, Any]] = []
    try:
        for sheet_name in SOURCE_SHEETS:
            sheet_records = _read_sheet(workbook[sheet_name], sheet_name)
            if len(sheet_records) != EXPECTED_SHEET_COUNTS[sheet_name]:
                raise OfficialReferenceImportError(
                    f"Unexpected record count for {sheet_name}: {len(sheet_records)}."
                )
            records.extend(sheet_records)
        _mark_hierarchy_reviews(records)
    finally:
        workbook.close()

    return {
        "asset_type": "official_reference_accounts",
        "schema_version": "1.0.0",
        "official_version_id": OFFICIAL_VERSION_ID,
        "base_status": "publicada",
        "approval_status": "aprovada",
        "methodology_version_id": METHODOLOGY_VERSION_ID,
        "description": (
            "Plano referencial oficial SPED/RFB para PJ em geral, "
            "ano-calendario 2024."
        ),
        "source_document_name": source_path.name,
        "source_document_hash": source_hash,
        "source_document_date": "2025-11-09",
        "source_url_or_reference": SOURCE_URL,
        "source_publisher": "Receita Federal do Brasil - SPED",
        "source_system": "ECF",
        "source_layout": "ECF_11",
        "source_calendar_year": SOURCE_CALENDAR_YEAR,
        "special_situations_year": 2025,
        "declaration_layout": "ECD_9",
        "entity_type": "PJ_GERAL",
        "source_sheets": list(SOURCE_SHEETS),
        "source_record_count": len(records),
        "asset_record_count": len(records),
        "coverage_status": "completa_com_revisao",
        "approval_notes": (
            "Fonte, escopo e cobertura liberados pelo usuario em 2026-07-24 "
            "para execucao no grupo."
        ),
        "required_fields": REQUIRED_FIELDS,
        "records": records,
    }


def _read_sheet(sheet: Any, sheet_name: str) -> list[dict[str, Any]]:
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    columns = {_normalize_header(value): index for index, value in enumerate(header)}
    required_columns = {
        "CODIGO",
        "DESCRICAO",
        "DT_INI",
        "DT_FIM",
        "TIPO",
        "CONTA SUPERIOR",
        "NIVEL",
        "NATUREZA",
        "ORIENTACOES",
    }
    missing = required_columns - set(columns)
    if missing:
        raise OfficialReferenceImportError(
            f"Missing columns in {sheet_name}: {', '.join(sorted(missing))}."
        )

    records = []
    for row in rows:
        code = _text(row[columns["CODIGO"]])
        if code is None:
            continue
        valid_from = _source_date(row[columns["DT_INI"]], "DT_INI", code)
        valid_to = _source_date(row[columns["DT_FIM"]], "DT_FIM", code, optional=True)
        nature_code = _required_text(row[columns["NATUREZA"]], "NATUREZA", code)
        nature = SOURCE_NATURES.get(nature_code)
        if nature is None:
            raise OfficialReferenceImportError(
                f"Unexpected nature {nature_code} for {code}."
            )
        source_type = _required_text(row[columns["TIPO"]], "TIPO", code)
        if source_type not in {"A", "S"}:
            raise OfficialReferenceImportError(
                f"Unexpected source type {source_type} for {code}."
            )

        records.append(
            {
                "reference_code": code,
                "official_description": _required_text(
                    row[columns["DESCRICAO"]], "DESCRICAO", code
                ),
                "parent_reference_code": _text(row[columns["CONTA SUPERIOR"]]),
                "level": _positive_int(row[columns["NIVEL"]], "NIVEL", code),
                "nature": nature,
                "valid_from": valid_from.year,
                "valid_to": valid_to.year if valid_to is not None else None,
                "layout": "ECD_9",
                "entity_type": "PJ_GERAL",
                "source": OFFICIAL_VERSION_ID,
                "status": _record_status(valid_to),
                "methodology_version_id": METHODOLOGY_VERSION_ID,
                "source_sheet": sheet_name,
                "source_type": source_type,
                "source_nature_code": nature_code,
                "source_valid_from": valid_from.isoformat(),
                "source_valid_to": valid_to.isoformat() if valid_to is not None else None,
                "official_guidance": _text(row[columns["ORIENTACOES"]]),
                "validation_notes": None,
            }
        )
    return records


def _mark_hierarchy_reviews(records: list[dict[str, Any]]) -> None:
    by_code = {record["reference_code"]: record for record in records}
    for record in records:
        parent_code = record["parent_reference_code"]
        if parent_code is None:
            continue
        parent = by_code.get(parent_code)
        if parent is None:
            raise OfficialReferenceImportError(
                f"Parent {parent_code} not found for {record['reference_code']}."
            )
        if parent["level"] >= record["level"]:
            record["status"] = "EM_REVISAO"
            record["validation_notes"] = (
                "Nivel oficial nao e superior ao nivel da conta pai; "
                "valor preservado sem correcao local."
            )


def _source_date(
    value: Any,
    field: str,
    code: str,
    *,
    optional: bool = False,
) -> date | None:
    if value is None or str(value).strip() == "":
        if optional:
            return None
        raise OfficialReferenceImportError(f"Missing {field} for {code}.")
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    try:
        return datetime.strptime(raw, "%d%m%Y").date()
    except ValueError as exc:
        raise OfficialReferenceImportError(
            f"Invalid {field} for {code}: {raw}."
        ) from exc


def _record_status(valid_to: date | None) -> str:
    calendar_year_end = date(SOURCE_CALENDAR_YEAR, 12, 31)
    return "INATIVA" if valid_to is not None and valid_to < calendar_year_end else "ATIVA"


def _positive_int(value: Any, field: str, code: str) -> int:
    if isinstance(value, bool):
        raise OfficialReferenceImportError(f"Invalid {field} for {code}.")
    try:
        parsed = int(value)
    except (TypeError, ValueError) as exc:
        raise OfficialReferenceImportError(f"Invalid {field} for {code}.") from exc
    if parsed < 1:
        raise OfficialReferenceImportError(f"Invalid {field} for {code}.")
    return parsed


def _required_text(value: Any, field: str, code: str) -> str:
    parsed = _text(value)
    if parsed is None:
        raise OfficialReferenceImportError(f"Missing {field} for {code}.")
    return parsed


def _text(value: Any) -> str | None:
    if value is None:
        return None
    parsed = str(value).strip()
    return parsed or None


def _normalize_header(value: Any) -> str:
    normalized = _required_text(value, "header", "header")
    return (
        normalized.upper()
        .replace("Á", "A")
        .replace("Ç", "C")
        .replace("Ã", "A")
        .replace("Õ", "O")
        .replace("Í", "I")
        .replace("Ó", "O")
        .replace("É", "E")
    )


if __name__ == "__main__":
    main()
